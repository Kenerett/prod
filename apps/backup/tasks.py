import os
import glob
import logging
from datetime import datetime, timedelta

from celery import shared_task
from django.conf import settings
from django.core.mail import EmailMessage
from django.utils import timezone

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def create_weekly_backup(self):
    """Creates weekly database backup with retry logic."""
    import subprocess
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"stms_backup_{timestamp}.sql"
        backup_path = os.path.join(str(settings.BACKUP_DIR), backup_filename)

        db = settings.DATABASES['default']
        cmd = ['pg_dump', '-h', db.get('HOST', ''), '-p', str(db.get('PORT', '5432')),
               '-U', db.get('USER', ''), '-d', db.get('NAME', ''), '-f', backup_path, '--verbose']

        env = os.environ.copy()
        env['PGPASSWORD'] = db.get('PASSWORD', '')

        result = subprocess.run(cmd, env=env, capture_output=True, text=True, timeout=600)

        if result.returncode != 0:
            raise RuntimeError(f"pg_dump failed: {result.stderr}")

        logger.info(f"Backup created: {backup_path}")
        excel_path = _create_excel_export(timestamp)
        _send_backup_notification(backup_path, excel_path)
        return f"Backup created: {backup_filename}"

    except Exception as exc:
        logger.error(f"Backup failed: {exc}")
        raise self.retry(exc=exc, countdown=60 * (2 ** self.request.retries))


def _create_excel_export(timestamp):
    from openpyxl import Workbook
    from school.models import (
        CustomUser, StudentProfile, TeacherAssignment,
        Grade, Attendance, Group, Subject, Semester,
        ScheduleEntry, TutorProfile
    )

    excel_path = os.path.join(str(settings.BACKUP_DIR), f"stms_data_{timestamp}.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    # Users sheet
    ws = wb.create_sheet("Users")
    ws.append(['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Middle Name', 'Role', 'Is Active', 'Date Joined'])
    for u in CustomUser.objects.all().iterator():
        ws.append([u.id, u.username, u.email, u.first_name, u.last_name,
                   u.middle_name or '', u.role, u.is_active,
                   u.date_joined.strftime('%Y-%m-%d %H:%M:%S')])

    # Students sheet
    ws = wb.create_sheet("Students")
    ws.append(['ID', 'User ID', 'Full Name', 'Groups'])
    for s in StudentProfile.objects.prefetch_related('groups').select_related('user').iterator():
        groups = ', '.join(g.name for g in s.groups.all())
        ws.append([s.id, s.user.id, s.user.get_full_name(), groups])

    # Groups sheet
    ws = wb.create_sheet("Groups")
    ws.append(['ID', 'Name', 'Students Count'])
    for g in Group.objects.all():
        ws.append([g.id, g.name, g.students.count()])

    # Grades sheet
    ws = wb.create_sheet("Grades")
    ws.append(['ID', 'Student', 'Subject', 'Teacher', 'Semester', 'Activity', 'Midterm', 'Final', 'Total'])
    for grade in Grade.objects.select_related(
        'student__user', 'teacher_assignment__teacher',
        'teacher_assignment__subject', 'semester'
    ).iterator():
        ws.append([
            grade.id, grade.student.user.get_full_name(),
            grade.teacher_assignment.subject.name,
            grade.teacher_assignment.teacher.get_full_name(),
            str(grade.semester), grade.activity or 0,
            grade.midterm or 0, grade.final or 0, grade.total or 0
        ])

    # Attendance sheet
    ws = wb.create_sheet("Attendance")
    ws.append(['ID', 'Student', 'Subject', 'Date', 'Missed Lessons', 'Reason'])
    for att in Attendance.objects.select_related(
        'student__user', 'teacher_assignment__subject'
    ).iterator():
        ws.append([
            att.id, att.student.user.get_full_name(),
            att.teacher_assignment.subject.name,
            att.date.strftime('%Y-%m-%d'), att.missed_lessons, att.reason or ''
        ])

    wb.save(excel_path)
    logger.info(f"Excel export created: {excel_path}")
    return excel_path


def _send_backup_notification(backup_path, excel_path):
    try:
        subject = f"STMS Weekly Backup - {datetime.now().strftime('%Y-%m-%d')}"
        body = (
            "Weekly backup of STMS completed successfully.\n\n"
            "Attached:\n- SQL database dump\n- Excel data export\n\n"
            "STMS System"
        )
        email = EmailMessage(
            subject=subject, body=body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[settings.ADMIN_EMAIL],
        )
        if os.path.exists(backup_path):
            email.attach_file(backup_path)
        if os.path.exists(excel_path):
            email.attach_file(excel_path)
        email.send()
        logger.info("Backup notification sent")
    except Exception as e:
        logger.error(f"Failed to send backup notification: {e}")


@shared_task
def cleanup_old_backups(days_to_keep=30):
    cutoff = datetime.now() - timedelta(days=days_to_keep)
    pattern = os.path.join(str(settings.BACKUP_DIR), "*")
    deleted = 0
    for path in glob.glob(pattern):
        if datetime.fromtimestamp(os.path.getctime(path)) < cutoff:
            try:
                os.remove(path)
                deleted += 1
                logger.info(f"Deleted old backup: {path}")
            except OSError as e:
                logger.error(f"Failed to delete {path}: {e}")
    return f"Deleted {deleted} old backup files"


@shared_task
def cleanup_old_logs(days=30):
    from school.models import RequestLog
    cutoff = timezone.now() - timedelta(days=days)
    deleted, _ = RequestLog.objects.filter(timestamp__lt=cutoff).delete()
    logger.info(f"Deleted {deleted} old log entries")
    return f"Deleted {deleted} old log entries"
