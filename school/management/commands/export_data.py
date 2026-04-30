# management/commands/export_data.py
from django.core.management.base import BaseCommand
from django.conf import settings
from datetime import datetime
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from school.models import (
    CustomUser, StudentProfile, Subject, Group, 
    TeacherAssignment, Grade, Semester
)

class Command(BaseCommand):
    help = 'Экспорт всех данных в Excel с email студентов'

    def add_arguments(self, parser):
        parser.add_argument(
            '--excel-only',
            action='store_true',
            help='Создать только Excel файл без SQL дампа',
        )
        parser.add_argument(
            '--output-dir',
            type=str,
            default=getattr(settings, 'BACKUP_DIR', 'exports/'),
            help='Директория для сохранения файлов',
        )
        parser.add_argument(
            '--group',
            type=str,
            help='Экспорт конкретной группы',
        )

    def handle(self, *args, **options):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = options['output_dir']
        group_name = options.get('group')
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        self.stdout.write(f"Создание экспорта данных в {output_dir}...")
        
        try:
            # Создание Excel файла
            excel_path = self.create_excel_export(timestamp, output_dir, group_name)
            self.stdout.write(
                self.style.SUCCESS(f'Excel файл создан: {excel_path}')
            )
            
            if not options['excel_only']:
                # Здесь можно добавить создание SQL дампа
                self.stdout.write("SQL дамп пропущен (используйте --excel-only для создания только Excel)")
            
            self.stdout.write(
                self.style.SUCCESS('Экспорт данных завершен успешно!')
            )
            
        except Exception as e:
            import traceback
            self.stdout.write(
                self.style.ERROR(f'Ошибка при экспорте: {str(e)}')
            )
            self.stdout.write(self.style.ERROR(traceback.format_exc()))

    def create_excel_export(self, timestamp, output_dir, group_name=None):
        """Создает Excel файл с данными студентов и оценками"""
        
        # Создаем workbook
        wb = Workbook()
        
        # Получаем группы для экспорта
        if group_name:
            groups = Group.objects.filter(name=group_name)
            if not groups.exists():
                raise ValueError(f"Группа '{group_name}' не найдена")
        else:
            groups = Group.objects.all()

        # Создаем лист для каждой группы
        for group in groups:
            self.stdout.write(f"Обрабатываем группу: {group.name}")
            self.create_group_sheet(wb, group)
        
        # Удаляем стандартный лист
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Сохраняем файл
        filename = f"students_export_{timestamp}.xlsx"
        if group_name:
            filename = f"students_export_{group_name}_{timestamp}.xlsx"
            
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)
        
        return filepath

    def create_group_sheet(self, workbook, group):
        """Создает лист с данными одной группы"""
        
        # Создаем лист с названием группы (ограничиваем длину)
        sheet_name = group.name[:31] if len(group.name) > 31 else group.name
        ws = workbook.create_sheet(title=sheet_name)
        
        # Получаем студентов группы
        student_profiles = StudentProfile.objects.filter(groups=group).select_related('user')
        
        if not student_profiles.exists():
            ws['A1'] = f"В группе {group.name} нет студентов"
            return
        
        # Получаем все предметы и семестры для группы
        assignments = TeacherAssignment.objects.filter(group=group).select_related(
            'subject', 'semester', 'teacher'
        ).order_by('semester__number', 'subject__name')
        
        # Создаем структуру заголовков
        self.create_headers(ws, assignments)
        
        # Заполняем данные студентов
        self.fill_student_data(ws, student_profiles, assignments, group)
        
        # Применяем стили
        self.apply_styles(ws, len(student_profiles), len(assignments))

    def create_headers(self, worksheet, assignments):
        """Создает заголовки таблицы"""
        
        # Группируем задания по семестрам
        semesters_data = {}
        for assignment in assignments:
            semester_num = assignment.semester.number
            if semester_num not in semesters_data:
                semesters_data[semester_num] = []
            semesters_data[semester_num].append(assignment)
        
        # Строка 1: Названия семестров
        current_col = 3  # Начинаем после Name, Email (без колонки кредиты в заголовке семестров)
        for semester_num in sorted(semesters_data.keys()):
            semester_assignments = semesters_data[semester_num]
            # Объединяем ячейки для названия семестра
            start_col = current_col
            end_col = current_col + len(semester_assignments) - 1
            
            if start_col == end_col:
                worksheet.cell(row=1, column=start_col, value=f"Семестр {semester_num}")
            else:
                worksheet.merge_cells(
                    start_row=1, start_column=start_col,
                    end_row=1, end_column=end_col
                )
                worksheet.cell(row=1, column=start_col, value=f"Семестр {semester_num}")
            
            current_col = end_col + 1
        
        # Строка 2: Кредиты по предметам
        worksheet.cell(row=2, column=1, value="Студент")
        worksheet.cell(row=2, column=2, value="Email")
        
        current_col = 3
        for assignment in assignments:
            # Получаем кредиты для предмета
            credits = assignment.subject.credits or ''
            worksheet.cell(row=2, column=current_col, value=credits)
            current_col += 1
        
        # Строка 3: Названия предметов с преподавателями
        worksheet.cell(row=3, column=1, value="")  # Пустая ячейка
        worksheet.cell(row=3, column=2, value="")  # Пустая ячейка
        
        current_col = 3
        for assignment in assignments:
            subject_name = assignment.subject.name
            teacher_name = f"{assignment.teacher.first_name} {assignment.teacher.last_name}"
            header = f"{subject_name} ({teacher_name})"
            worksheet.cell(row=3, column=current_col, value=header)
            current_col += 1

    def fill_student_data(self, worksheet, student_profiles, assignments, group):
        """Заполняет данные студентов"""
        
        row = 4  # Начинаем с четвертой строки (после 3 строк заголовков)
        
        for profile in student_profiles.order_by('user__last_name', 'user__first_name'):
            user = profile.user
            
            # ФИО студента
            full_name = f"{user.last_name} {user.first_name}"
            if user.middle_name:
                full_name += f" {user.middle_name}"
            
            worksheet.cell(row=row, column=1, value=full_name)
            
            # Email студента
            worksheet.cell(row=row, column=2, value=user.email or "")
            
            # Заполняем оценки (начинаем с колонки 3, так как убрали отдельную колонку кредитов)
            current_col = 3
            for assignment in assignments:
                grade = Grade.objects.filter(
                    student=profile,
                    teacher_assignment=assignment
                ).first()
                
                grade_value = ""
                if grade:
                    if grade.total:
                        grade_value = grade.total
                    elif grade.additional_scores and assignment.subject.name in grade.additional_scores:
                        grade_value = grade.additional_scores[assignment.subject.name]
                
                worksheet.cell(row=row, column=current_col, value=grade_value)
                current_col += 1
            
            row += 1
            
            # Выводим прогресс
            if row % 10 == 0:
                self.stdout.write(f"  Обработано {row-3} студентов...")  # -3 потому что начинаем с 4 строки

    def apply_styles(self, worksheet, num_students, num_assignments):
        """Применяет стили к таблице"""
        
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Применяем стили к заголовкам семестров (строка 1)
        for col in range(1, 3 + num_assignments):
            cell = worksheet.cell(row=1, column=col)
            if cell.value:  # Только если есть значение
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
        
        # Применяем стили к строке с кредитами (строка 2)
        for col in range(1, 3 + num_assignments):
            cell = worksheet.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Применяем стили к заголовкам предметов (строка 3)
        for col in range(1, 3 + num_assignments):
            cell = worksheet.cell(row=3, column=col)
            cell.font = header_font
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Применяем границы к данным
        for row in range(4, 4 + num_students):  # Начинаем с 4 строки
            for col in range(1, 3 + num_assignments):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                
                # Выравнивание для разных колонок
                if col == 1:  # ФИО
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col == 2:  # Email
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:  # Оценки
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Автоширина колонок
        for col in range(1, 3 + num_assignments):
            column_letter = worksheet.cell(row=1, column=col).column_letter
            
            if col == 1:  # ФИО - широкая колонка
                worksheet.column_dimensions[column_letter].width = 25
            elif col == 2:  # Email - широкая колонка
                worksheet.column_dimensions[column_letter].width = 30
            else:  # Оценки
                worksheet.column_dimensions[column_letter].width = 15
        
        # Фиксируем первые строки (теперь 3 строки заголовков)
        worksheet.freeze_panes = "A4"