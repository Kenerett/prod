import os
import subprocess
from datetime import datetime
from django.conf import settings
from django.core.mail import EmailMessage
from celery import shared_task
import logging
import glob
from datetime import timedelta

logger = logging.getLogger(__name__)

@shared_task
def create_weekly_backup():
    """Создает еженедельное резервное копирование базы данных"""
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"stms_backup_{timestamp}.sql"
        backup_path = os.path.join(settings.BACKUP_DIR, backup_filename)
        
        # Команда для создания дампа PostgreSQL
        cmd = [
            'pg_dump',
            '-h', settings.DATABASES['default']['HOST'],
            '-p', settings.DATABASES['default']['PORT'],
            '-U', settings.DATABASES['default']['USER'],
            '-d', settings.DATABASES['default']['NAME'],
            '-f', backup_path,
            '--verbose'
        ]
        
        # Установка переменной окружения для пароля
        env = os.environ.copy()
        env['PGPASSWORD'] = settings.DATABASES['default']['PASSWORD']
        
        # Выполнение команды
        result = subprocess.run(cmd, env=env, capture_output=True, text=True)
        
        if result.returncode == 0:
            logger.info(f"Резервная копия успешно создана: {backup_path}")
            
            # Создание Excel файла с данными
            excel_path = create_excel_export(timestamp)
            
            # Отправка уведомления администратору (опционально)
            send_backup_notification(backup_path, excel_path)
            
            return f"Backup created successfully: {backup_filename}"
        else:
            logger.error(f"Ошибка создания резервной копии: {result.stderr}")
            return f"Backup failed: {result.stderr}"
            
    except Exception as e:
        logger.error(f"Исключение при создании резервной копии: {str(e)}")
        return f"Backup failed with exception: {str(e)}"

def create_excel_export(timestamp):
    """Создает Excel файл со всеми данными системы"""
    try:
        from openpyxl import Workbook
        from school.models import (CustomUser, StudentProfile, TeacherAssignment, 
                            Grade, Attendance, Group, Subject, Semester, 
                            ScheduleEntry, TutorProfile)
        
        excel_filename = f"stms_data_{timestamp}.xlsx"
        excel_path = os.path.join(settings.BACKUP_DIR, excel_filename)
        
        wb = Workbook()
        
        # Удаляем стандартный лист
        wb.remove(wb.active)
        
        # Экспорт пользователей
        users_ws = wb.create_sheet("Users")
        users_ws.append([
            'ID', 'Username', 'Email', 'First Name', 'Last Name', 'Middle Name', 
            'Role', 'Is Active', 'Date Joined'
        ])
        
        for user in CustomUser.objects.all():
            users_ws.append([
                user.id, user.username, user.email, user.first_name, 
                user.last_name, user.middle_name or '', user.role, 
                user.is_active, user.date_joined.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        # Экспорт студентов
        students_ws = wb.create_sheet("Students")
        students_ws.append([
            'ID', 'User ID', 'Full Name', 'Groups', 'Current Semester'
        ])
        
        for student in StudentProfile.objects.all():
            students_ws.append([
                student.id, student.user.id, student.user.get_full_name(),
                student.get_all_groups(), 
                str(student.get_current_semester()) if student.get_current_semester() else 'N/A'
            ])
        
        # Экспорт групп
        groups_ws = wb.create_sheet("Groups")
        groups_ws.append(['ID', 'Name', 'Students Count'])
        
        for group in Group.objects.all():
            groups_ws.append([
                group.id, group.name, group.students.count()
            ])
        
        # Экспорт предметов
        subjects_ws = wb.create_sheet("Subjects")
        subjects_ws.append(['ID', 'Name', 'Description', 'Credits'])
        
        for subject in Subject.objects.all():
            subjects_ws.append([
                subject.id, subject.name, subject.description or '', subject.credits
            ])
        
        # Экспорт назначений учителей
        assignments_ws = wb.create_sheet("Teacher Assignments")
        assignments_ws.append([
            'ID', 'Teacher', 'Group', 'Subject', 'Semester', 'Num SG'
        ])
        
        for assignment in TeacherAssignment.objects.all():
            assignments_ws.append([
                assignment.id, assignment.teacher.get_full_name(),
                assignment.group.name, assignment.subject.name,
                str(assignment.semester), assignment.num_sg or 0
            ])
        
        # Экспорт оценок
        grades_ws = wb.create_sheet("Grades")
        grades_ws.append([
            'ID', 'Student', 'Teacher Assignment', 'Semester', 
            'Activity', 'Midterm', 'Final', 'Total', 'SG Scores'
        ])
        
        for grade in Grade.objects.all():
            sg_scores = ', '.join([f"{k}: {v}" for k, v in grade.get_sg_scores().items()])
            grades_ws.append([
                grade.id, grade.student.user.get_full_name(),
                f"{grade.teacher_assignment.teacher.get_full_name()} - {grade.teacher_assignment.subject.name}",
                str(grade.semester), grade.activity or 0, grade.midterm or 0,
                grade.final or 0, grade.total or 0, sg_scores
            ])
        
        # Экспорт посещаемости
        attendance_ws = wb.create_sheet("Attendance")
        attendance_ws.append([
            'ID', 'Student', 'Teacher Assignment', 'Date', 'Missed Lessons', 'Reason'
        ])
        
        for attendance in Attendance.objects.all():
            attendance_ws.append([
                attendance.id, attendance.student.user.get_full_name(),
                f"{attendance.teacher_assignment.teacher.get_full_name()} - {attendance.teacher_assignment.subject.name}",
                attendance.date.strftime('%Y-%m-%d'), attendance.missed_lessons,
                attendance.reason or ''
            ])
        
        # Экспорт расписания
        schedule_ws = wb.create_sheet("Schedule")
        schedule_ws.append([
            'ID', 'Scheduler', 'Weekday', 'Week Type', 'Time Slot', 
            'Group', 'Teacher', 'Subject', 'Room'
        ])
        
        for entry in ScheduleEntry.objects.all():
            schedule_ws.append([
                entry.id, entry.scheduler.get_full_name(),
                entry.get_weekday_display(), entry.get_week_type_display(),
                entry.get_time_slot_display(), entry.group.name,
                entry.teacher.get_full_name(), entry.subject.name, str(entry.room)
            ])
        
        # Сохранение файла
        wb.save(excel_path)
        logger.info(f"Excel файл создан: {excel_path}")
        
        return excel_path
        
    except Exception as e:
        logger.error(f"Ошибка создания Excel файла: {str(e)}")
        raise

def send_backup_notification(backup_path, excel_path):
    """Отправляет уведомление администратору о создании резервной копии"""
    try:
        subject = f"Еженедельная резервная копия STMS - {datetime.now().strftime('%Y-%m-%d')}"
        message = """
        Здравствуйте!
        
        Еженедельная резервная копия системы STMS была успешно создана.
        
        В приложении:
        - SQL дамп базы данных
        - Excel файл со всеми данными системы
        
        С уважением,
        Система STMS
        """
        
        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[settings.ADMIN_EMAIL],
        )
        
        # Прикрепление файлов
        if os.path.exists(backup_path):
            email.attach_file(backup_path)
        if os.path.exists(excel_path):
            email.attach_file(excel_path)
        
        email.send()
        logger.info("Уведомление о резервной копии отправлено")
        
    except Exception as e:
        logger.error(f"Ошибка отправки уведомления: {str(e)}")

@shared_task
def cleanup_old_backups(days_to_keep=30):
    """Удаляет старые резервные копии"""
    try:
        cutoff_date = datetime.now() - timedelta(days=days_to_keep)
        
        pattern = os.path.join(settings.BACKUP_DIR, "*")
        files = glob.glob(pattern)
        
        deleted_count = 0
        for file_path in files:
            file_time = datetime.fromtimestamp(os.path.getctime(file_path))
            if file_time < cutoff_date:
                try:
                    os.remove(file_path)
                    deleted_count += 1
                    logger.info(f"Удален старый файл резервной копии: {file_path}")
                except Exception as e:
                    logger.error(f"Ошибка удаления файла {file_path}: {str(e)}")
        
        return f"Удалено {deleted_count} старых файлов резервных копий"
        
    except Exception as e:
        logger.error(f"Ошибка очистки старых файлов: {str(e)}")
        return f"Ошибка: {str(e)}"


from celery import shared_task
from django.utils import timezone
from .models import RequestLog

@shared_task
def cleanup_old_logs(days=30):
    cutoff = timezone.now() - timezone.timedelta(days=days)
    deleted, _ = RequestLog.objects.filter(timestamp__lt=cutoff).delete()
    return f"Удалено {deleted} старых логов"