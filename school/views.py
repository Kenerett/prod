# school/views.py
import logging
import os
import sys
import json
import tempfile
from collections import defaultdict
from datetime import datetime, timedelta
from io import StringIO
from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import SetPasswordForm
from django.contrib.auth.tokens import default_token_generator
from django.contrib.sites.shortcuts import get_current_site
from django.core.exceptions import ValidationError, PermissionDenied
from django.core.mail import send_mail
from django.core.management import call_command
from django.db import models, transaction
from django.db.models import Avg, Count, Q, Sum, Prefetch
from django.http import HttpResponse, Http404, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.views.decorators.http import require_http_methods
from django.views.generic import TemplateView
from .forms import ScheduleEntryForm, GroupScheduleForm
import pandas as pd
import openpyxl
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode, url_has_allowed_host_and_scheme



from .forms import (
    ScheduleEntryForm, 
    ImportExcelForm, 
    CustomUserCreationForm,
)
from .models import (
    StudentProfile, Attendance, Grade, Room, ScheduleEntry,
    Material, CustomUser, TeacherAssignment, Group, Subject, 
    TutorProfile, GlobalGradeSettings, Semester
)
from .utils.word_scanner import WordSubjectScanner # Убедитесь, что этот модуль существует и корректен

logger = logging.getLogger(__name__)


def is_tutor(user):
    return user.is_authenticated and hasattr(user, 'role') and user.role == CustomUser.TUTOR


WEEKDAYS = [
    (1, 'Понедельник'),
    (2, 'Вторник'),
    (3, 'Среда'),
    (4, 'Четверг'),
    (5, 'Пятница'),
]

from evaluation.models import Evaluation, EvaluationSettings
from school.models import TeacherAssignment, Semester



@login_required
def student_dashboard(request):
    # Автоматическая проверка оценки для студентов

    # Основной код dashboard'а
# Основной код dashboard'а
    try:
        student_profile = request.user.student_profile
    except StudentProfile.DoesNotExist:
        messages.error(request, 'Профиль студента не найден.')
        return redirect('logout')

    student_groups = student_profile.groups.all()

    group_assignments = TeacherAssignment.objects.filter(
        group__in=student_groups
    ).select_related('teacher', 'group', 'subject', 'semester')

    # Один запрос вместо N
    grades_map = {
        g.teacher_assignment_id: g
        for g in Grade.objects.filter(
            student=student_profile,
            teacher_assignment__in=group_assignments,
        )
    }

    subjects_by_semester = {}
    semester_gpa = {}

    for assignment in group_assignments:
        semester = assignment.semester
        if semester not in subjects_by_semester:
            subjects_by_semester[semester] = []

        grade = grades_map.get(assignment.id)  # без SQL
        
        # Формируем данные предмета
        subject_data = {
            'id': assignment.subject.id,
            'name': assignment.subject.name,
            'teacher': assignment.teacher.get_full_name(),
            'group': assignment.group.name,
            'assignment_id': assignment.id,
            'credits': assignment.subject.credits,  # Добавляем количество кредитов
            'grades': []
        }
        
        if grade:
            sg_scores = grade.get_sg_scores()
            subject_data['grades'].append({
                'activity': grade.activity,
                'sg_scores': sg_scores,
                'sg_total': grade.calculate_sg_total(),
                'midterm': grade.midterm,
                'final': grade.final,
                'total': grade.total
            })
        
        # Добавляем предмет в соответствующий семестр
        subjects_by_semester[semester].append(subject_data)

    # Вычисляем GPA для каждого семестра по правильной формуле
    for semester, subjects in subjects_by_semester.items():
        total_weighted_score = 0
        total_credits = 0
        
        # Собираем оценки для этого семестра
        semester_grades = []
        for subject in subjects:
            for grade_data in subject['grades']:
                if grade_data['total'] is not None:
                    semester_grades.append({
                        'total': grade_data['total'],
                        'credits': subject['credits']
                    })
        
        # Вычисляем взвешенное среднее
        for grade_info in semester_grades:
            if grade_info['credits'] > 0:
                total_weighted_score += grade_info['total'] * grade_info['credits']
                total_credits += grade_info['credits']
        
        if total_credits > 0:
            semester_gpa[semester.number] = round(total_weighted_score / total_credits, 2)

    # Общий GPA (среднее по всем семестрам) по правильной формуле
    total_weighted_score = 0
    total_credits = 0
    
    # Собираем все оценки
    all_grades = []
    for semester, subjects in subjects_by_semester.items():
        for subject in subjects:
            for grade_data in subject['grades']:
                if grade_data['total'] is not None:
                    all_grades.append({
                        'total': grade_data['total'],
                        'credits': subject['credits']
                    })
    
    # Вычисляем взвешенное среднее
    for grade_info in all_grades:
        if grade_info['credits'] > 0:
            total_weighted_score += grade_info['total'] * grade_info['credits']
            total_credits += grade_info['credits']
    
    gpa = round(total_weighted_score / total_credits, 2) if total_credits > 0 else None

    # Посещаемость (оставляем без изменений)
    attendance_qs = Attendance.objects.filter(student=student_profile).select_related('teacher_assignment__subject')
    attendance_data = {}
    for att in attendance_qs:
        subject_name = att.teacher_assignment.subject.name
        attendance_data[subject_name] = attendance_data.get(subject_name, 0) + att.missed_lessons

    # Учебные материалы (оставляем без изменений)
    presentations = Material.objects.all().select_related('teacher_assignment__subject').order_by('-uploaded_at')

    # Расписание (оставляем без изменений)
    schedule_data = {}
    if student_groups.exists():
        schedule_entries = ScheduleEntry.objects.filter(
            group__in=student_groups
        ).select_related('teacher', 'subject', 'room').order_by('week_type', 'weekday', 'time_slot')
        
        for entry in schedule_entries:
            week_key = entry.week_type
            weekday = entry.weekday
            
            if week_key not in schedule_data:
                schedule_data[week_key] = {}
            
            if weekday not in schedule_data[week_key]:
                schedule_data[week_key][weekday] = []
            
            start_time, end_time = entry.start_end_time
            
            schedule_data[week_key][weekday].append({
                'time': f"{start_time}-{end_time}",
                'subject': entry.subject.name,
                'teacher': entry.teacher.get_full_name(),
                'room': str(entry.room),
                'group': str(entry.group),
                'time_slot': entry.time_slot
            })
        
        for week_type in schedule_data:
            for weekday in schedule_data[week_type]:
                schedule_data[week_type][weekday].sort(key=lambda x: x['time_slot'])

    # Словари названий
    weekday_names = {
        0: 'Понедельник',
        1: 'Вторник', 
        2: 'Среда',
        3: 'Четверг',
        4: 'Пятница'
    }

    week_names = {
        'top': 'Верхняя неделя',
        'bottom': 'Нижняя неделя'
    }

    return render(request, 'school/student_dashboard.html', {
        'student': student_profile,
        'subjects_by_semester': subjects_by_semester,
        'semester_gpa': semester_gpa,
        'attendance_data': attendance_data,
        'gpa': gpa,
        'presentations': presentations,
        'schedule_data': schedule_data,
        'weekday_names': weekday_names,
        'week_names': week_names,
    })


@login_required
def student_subject_detail(request, tsg_id):
    # Используем related_name для получения профиля
    student_profile = getattr(request.user, 'student_profile', None)
    if not student_profile:
         messages.error(request, 'Профиль студента не найден.')
         return redirect('logout')

    tsg = get_object_or_404(TeacherAssignment, id=tsg_id)

    if tsg.group not in student_profile.groups.all():
        messages.error(request, 'Доступ запрещён.')
        return redirect('student_dashboard')

    grade = Grade.objects.filter(student=student_profile, teacher_assignment=tsg).first()
    
    grade_data = None
    if grade:
        sg_scores = grade.get_sg_scores()
        grade_data = {
            'activity': grade.activity,
            'sg_scores': sg_scores,
            'sg_total': grade.calculate_sg_total(),
            'midterm': grade.midterm,
            'final': grade.final,
            'total': grade.total
        }

    attendance = Attendance.objects.filter(student=student_profile, teacher_assignment=tsg).order_by('-date')
    materials = Material.objects.filter(teacher_assignment=tsg).order_by('-uploaded_at')

    return render(request, 'school/student_subject_detail.html', {
        'tsg': tsg,
        'grade': grade_data,
        'attendance': attendance,
        'materials': materials,
    })


def get_time_range_from_slot(time_slot):
    """
    Преобразует time_slot (int) в строку времени начала и окончания.
    Возвращает кортеж (start_time_str, end_time_str).
    """
    TIME_SLOT_MAPPING = {
        1: ("08:30", "09:50"),
        2: ("10:05", "11:25"),
        3: ("11:40", "13:00"),
        4: ("13:30", "14:50"),
        5: ("15:05", "16:25"),
        6: ("16:40", "18:00"),
        7: ("18:15", "19:35"),
        8: ("19:50", "21:10"),
    }
    
    if time_slot in TIME_SLOT_MAPPING:
        return TIME_SLOT_MAPPING[time_slot]
    else:
        # Возвращаем дефолтное значение для неизвестных слотов
        return (f"Время {time_slot}", f"Время {time_slot}")


from django.shortcuts import render


@login_required
def teacher_dashboard(request):
    # Получаем назначения преподавателя
    assignments = TeacherAssignment.objects.filter(teacher=request.user).select_related('group', 'subject', 'semester')
    
    # Получаем уникальные группы для статистики
    total_groups = set(assignment.group for assignment in assignments)
    
    # Получаем расписание преподавателя
    schedule_data = {}
    if assignments.exists():
        # Получаем все записи расписания для этого преподавателя
        schedule_entries = ScheduleEntry.objects.filter(
            teacher=request.user
        ).select_related('group', 'subject', 'room').order_by('week_type', 'weekday', 'time_slot')
        
        # Группируем по типу недели
        for entry in schedule_entries:
            week_key = entry.week_type
            
            if week_key not in schedule_data:
                schedule_data[week_key] = []
            
            # Добавляем информацию о занятии
            schedule_data[week_key].append({
                'weekday': entry.weekday,
                'time_slot': entry.time_slot,
                'subject': entry.subject.name,
                'group': entry.group.name,
                'room': str(entry.room),
                'time': f"{entry.start_end_time[0]}-{entry.start_end_time[1]}"
            })

    # Словари названий для шаблона
    weekday_names = {
        0: 'Понедельник',
        1: 'Вторник', 
        2: 'Среда',
        3: 'Четверг',
        4: 'Пятница'
    }

    week_names = {
        'top': 'Верхняя неделя',
        'bottom': 'Нижняя неделя'
    }

    # Временные слоты для отображения
    time_slots = [
        (1, '8:30-9:50'),
        (2, '10:05-11:25'),
        (3, '11:40-13:00'),
        (4, '13:30-14:50'),
        (5, '15:05-16:25'),
        (6, '16:40-18:00'),
    ]

    return render(request, 'school/teacher_dashboard.html', {
        'teacher_subjects': assignments,
        'total_groups': total_groups,
        'schedule_data': schedule_data,
        'weekday_names': weekday_names,
        'week_names': week_names,
        'time_slots': time_slots,
    })








# views.py
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import render, redirect
from django.utils import timezone
from datetime import timedelta
from .models import CustomUser
from axes.decorators import axes_dispatch
from axes.utils import reset
from django.utils.http import url_has_allowed_host_and_scheme

# views.py
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import render, redirect
from .models import CustomUser
from axes.decorators import axes_dispatch
from axes.utils import reset
from django.utils.decorators import method_decorator
@axes_dispatch
def user_login(request):
    context = {}
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Проверяем, не заблокирован ли IP через Axes
        # Axes автоматически проверит это и покажет страницу блокировки если нужно
        
        user = CustomUser.objects.filter(username=username).first()
        if user:
            user_auth = authenticate(request, username=username, password=password)
            if user_auth is not None:
                # Сброс счетчика Axes при успешном входе
                
                login(request, user_auth)
                
                # Перенаправление в зависимости от роли
                next_url = request.POST.get('next') or request.GET.get('next')
                if next_url and url_has_allowed_host_and_scheme(url=next_url, allowed_hosts={request.get_host()}):
                    return redirect(next_url)
                    
                if user.role == 'student':
                    return redirect('student_dashboard')
                elif user.role == 'teacher':
                    return redirect('teacher_dashboard')
                elif user.role == 'tutor':
                    return redirect('tutor_dashboard')
                elif user.role == 'scheduler':
                    return redirect('scheduler_dashboard')
                elif user.is_superuser:
                    return redirect('admin_dashboard')
                else:
                    return redirect('dashboard')
            else:
                # При неудачной попытке Axes автоматически увеличит счетчик
                messages.error(request, 'Incorrect username or password.')
        else:
            # Даже для несуществующих пользователей Axes будет учитывать попытку
            messages.error(request, 'Incorrect username or password.')
        context['username'] = username
    return render(request, 'school/login.html', context)






def user_logout(request):
    logout(request)
    return redirect('login')

@login_required
def dashboard(request):
    role = getattr(request.user, 'role', None)
    if role == 'student':
        return redirect('student_dashboard')
    elif role == 'teacher':
        return redirect('teacher_dashboard')
    elif request.user.is_superuser:
        return redirect('admin_dashboard')
    else:
        return redirect('login')




@login_required
def grade_students(request, tsg_id):
    """
    View for teachers to manage grades and attendance for a specific Teacher-Subject-Group assignment.
    Handles setting the number of quizzes (SG) and entering grades/attendance.
    Implements logic for midterm/final limits set by admin.
    """
    # --- Authentication and Authorization ---
    if request.user.role != 'teacher':
        messages.error(request, 'Access denied.')
        return redirect('logout')

    # --- Get the TeacherAssignment ---
    tsg = get_object_or_404(
        TeacherAssignment.objects.select_related('teacher', 'group', 'subject', 'semester'), 
        id=tsg_id
    )
    # Ensure the logged-in teacher owns this assignment
    if tsg.teacher != request.user:
        messages.error(request, 'You cannot manage grades for this group/subject.')
        return redirect('teacher_dashboard')

    students = tsg.group.students.all().select_related('user')
    
    # --- Step 1: Set Number of SG (Quizzes) ---
    if tsg.num_sg is None:
        if request.method == 'POST':
            try:
                num_sg_str = request.POST.get('num_sg')
                if not num_sg_str:
                    raise ValueError("Number not specified")
                num_sg = int(num_sg_str)
                if num_sg < 0:
                    raise ValueError("Number cannot be negative")
                if num_sg > 20:
                    raise ValueError("Number of quizzes cannot exceed 20")
                tsg.num_sg = num_sg
                tsg.save(update_fields=['num_sg'])
                messages.success(request, f'Number of quizzes (SG) set: {num_sg}')
                return redirect('grade_students', tsg_id=tsg.id)
            except (ValueError, TypeError) as e:
                messages.error(request, f'Error entering number of quizzes: {e}')
        
        context = {
            'tsg': tsg,
            'students': students,
            'setting_num_sg': True
        }
        return render(request, 'school/grade_form.html', context)

    # --- Step 2: Main Grading and Attendance Form (if num_sg is set) ---
    if request.method == 'POST':
        
        # --- Handle Material Upload ---
        if 'upload_material' in request.POST:
            title = request.POST.get('material_title')
            uploaded_file = request.FILES.get('material_file')
            if title and uploaded_file:
                try:
                    Material.objects.create(
                        teacher_assignment=tsg,
                        title=title,
                        file=uploaded_file
                    )
                    messages.success(request, 'Material uploaded successfully!')
                except Exception as e:
                    messages.error(request, f'Error uploading material: {e}')
            else:
                messages.error(request, 'Please specify a title and select a file to upload material.')
            return redirect('grade_students', tsg_id=tsg.id)

        # --- Handle Grades and Attendance Submission ---
        missed_date_str = request.POST.get('common_date')
        all_students_processed = True
        errors_for_alert = []

        # Load global grade limits
        grade_settings = GlobalGradeSettings.load()
        existing_grades_map = {
            g.student_id: g
            for g in Grade.objects.filter(teacher_assignment=tsg)
        }
        for student in students:
            student_full_name = student.user.get_full_name() if student.user else f"Student ID {student.id}"

            def parse_grade_field(field_name_post_key):
                value_str = request.POST.get(field_name_post_key, '').strip()
                if value_str == '':
                    return None
                try:
                    return float(value_str) if '.' in value_str else int(value_str)
                except (ValueError, TypeError):
                    messages.warning(
                        request, 
                        f'Invalid value for field \'{field_name_post_key}\' for {student_full_name}: \'{value_str}\''
                    )
                    return None

            # --- Midterm grade logic with limit check ---
            existing_grade = existing_grades_map.get(student.id)

            midterm_new = parse_grade_field(f'midterm_{student.id}')
            
            if grade_settings.midterm_limit is None: # Limit disabled
                if existing_grade and existing_grade.midterm is not None:
                    # If teacher tries to change existing value
                    if midterm_new is not None and midterm_new != existing_grade.midterm:
                        messages.error(
                            request,
                            f'Changing "Midterm" grade for student {student_full_name} is forbidden by admin (limit disabled).'
                        )
                        all_students_processed = False
                        continue
                    else:
                        # Keep old value
                        midterm_to_save = existing_grade.midterm
                else:
                    # No grade and limit disabled — don't save new grade
                    if midterm_new is not None:
                        messages.error(
                            request,
                            f'Setting "Midterm" grade for student {student_full_name} is forbidden by admin (limit disabled).'
                        )
                        all_students_processed = False
                        continue
                    midterm_to_save = None # Explicitly set None
            else: # Limit enabled
                # Check value within limit
                if midterm_new is not None and (midterm_new < 0 or midterm_new > grade_settings.midterm_limit):
                     messages.error(
                        request,
                        f'"Midterm" grade for student {student_full_name} must be between 0 and {grade_settings.midterm_limit} (limit set by admin).'
                    )
                     all_students_processed = False
                     continue
                midterm_to_save = midterm_new # Save new value or None

            # --- Final grade logic with limit check ---
            final_new = parse_grade_field(f'final_{student.id}')
            
            if grade_settings.final_limit is None: # Limit disabled
                if existing_grade and existing_grade.final is not None:
                    # If teacher tries to change existing value
                    if final_new is not None and final_new != existing_grade.final:
                        messages.error(
                            request,
                            f'Changing "Final" grade for student {student_full_name} is forbidden by admin (limit disabled).'
                        )
                        all_students_processed = False
                        continue
                    else:
                        # Keep old value
                        final_to_save = existing_grade.final
                else:
                    # No grade and limit disabled — don't save new grade
                    if final_new is not None:
                        messages.error(
                            request,
                            f'Setting "Final" grade for student {student_full_name} is forbidden by admin (limit disabled).'
                        )
                        all_students_processed = False
                        continue
                    final_to_save = None # Explicitly set None
            else: # Limit enabled
                # Check value within limit
                if final_new is not None and (final_new < 0 or final_new > grade_settings.final_limit):
                     messages.error(
                        request,
                        f'"Final" grade for student {student_full_name} must be between 0 and {grade_settings.final_limit} (limit set by admin).'
                    )
                     all_students_processed = False
                     continue
                final_to_save = final_new # Save new value or None

            # --- Handle activity ---
            activity = parse_grade_field(f'activity_{student.id}')

            # --- Parse SG (Quiz) Fields ---
            sg_scores = {}
            total_sg_temp = 0.0
            expected_sg_fields = tsg.get_sg_field_names()
            for sg_key in expected_sg_fields:
                field_name_post_key = f'sg_{sg_key}_{student.id}'
                value_str = request.POST.get(field_name_post_key, '').strip()
                if value_str == '':
                    sg_scores[sg_key] = 0
                else:
                    try:
                        score_val = float(value_str) if '.' in value_str else int(value_str)
                        sg_scores[sg_key] = score_val
                        total_sg_temp += score_val
                    except (ValueError, TypeError):
                        messages.warning(
                            request, 
                            f'Invalid value {sg_key} for student {student_full_name}: \'{value_str}\''
                        )
                        sg_scores[sg_key] = 0
            
            # --- Server-Side Validation for SG Total ---
            if total_sg_temp > 20.0:
                error_msg = f'Error for {student_full_name}: SG total ({total_sg_temp:.1f}) exceeds 20.'
                messages.error(request, error_msg)
                errors_for_alert.append(f"{student_full_name}: SG total = {total_sg_temp:.1f} > 20")
                all_students_processed = False
                continue

            # --- Update or Create Grade Object ---
            grade, created = Grade.objects.get_or_create(
                student=student,
                teacher_assignment=tsg,
                defaults={
                    'activity': activity,
                    'midterm': midterm_to_save,
                    'final': final_to_save,
                    'additional_scores': sg_scores,
                    'semester': tsg.semester
                }
            )
            
            if not created and grade.semester is None:
                grade.semester = tsg.semester

            grade_updated = False
            if (activity is not None and grade.activity != activity) or \
               (midterm_to_save is not None and grade.midterm != midterm_to_save) or \
               (final_to_save is not None and grade.final != final_to_save) or \
               (grade.additional_scores != sg_scores) or \
               (grade.semester != tsg.semester):
                
                if activity is not None: grade.activity = activity
                if midterm_to_save is not None: grade.midterm = midterm_to_save
                if final_to_save is not None: grade.final = final_to_save
                if grade.additional_scores != sg_scores: grade.additional_scores = sg_scores
                if grade.semester != tsg.semester: grade.semester = tsg.semester
                grade_updated = True

            if grade_updated:
                try:
                    grade.save(update_fields=['activity', 'midterm', 'final', 'additional_scores', 'total', 'semester'])
                    logger.info(f"Grades updated for student {student_full_name} (ID: {student.id})")
                except ValidationError as e:
                    messages.error(request, f'Grade validation error for {student_full_name}: {e}')
                    logger.error(f"ValidationError when saving grades for {student_full_name}: {e}", exc_info=True)
                    all_students_processed = False
                except Exception as e:
                    messages.error(request, f'Error saving grades for {student_full_name}: {e}')
                    logger.error(f"Unexpected error saving grades for {student_full_name}: {e}", exc_info=True)
                    all_students_processed = False

            # --- Handle Attendance Submission ---
            if missed_date_str:
                try:
                    date_obj = datetime.strptime(missed_date_str, '%Y-%m-%d').date()
                    
                    for student in students:
                        student_full_name = student.user.get_full_name() if student.user else f"Student ID {student.id}"
                        # Get missed lessons count from form for each student
                        missed_lessons_str = request.POST.get(f'missed_lessons_{student.id}', '0').strip()
                        try:
                            missed_lessons_val = int(missed_lessons_str) if missed_lessons_str else 0
                            if missed_lessons_val < 0: 
                                missed_lessons_val = 0
                        except ValueError:
                            messages.warning(request, f'Invalid missed lessons value for {student_full_name}. Set to 0.')
                            missed_lessons_val = 0
                        
                        # Reason is saved only if there are missed lessons
                        reason_val = ''
                        if missed_lessons_val > 0:
                            # If reason field is missing from template, leave empty
                            reason_val = '' 

                        try:
                            # Update or create attendance record
                            Attendance.objects.update_or_create(
                                student=student,
                                teacher_assignment=tsg,
                                date=date_obj,
                                defaults={
                                    'missed_lessons': missed_lessons_val,
                                    'reason': reason_val if missed_lessons_val > 0 else ''
                                }
                            )
                        except Exception as e:
                            error_msg = f'Error processing attendance for {student_full_name}: {e}'
                            messages.error(request, error_msg)
                            logger.error(error_msg, exc_info=True)
                            all_students_processed = False

                except ValueError:
                    error_msg = f'Invalid date format: {missed_date_str}'
                    messages.error(request, error_msg)
                    logger.error(error_msg)
                    all_students_processed = False
                except Exception as e:
                    error_msg = f'Error processing attendance: {e}'
                    messages.error(request, error_msg)
                    logger.error(error_msg, exc_info=True)
                    all_students_processed = False


        # --- Final Feedback ---
        if all_students_processed:
            messages.success(request, 'Grades and attendance updated successfully!')
        else:
            if not errors_for_alert:
                 messages.warning(request, 'Update completed with errors. Check messages above.')

        return redirect('grade_students', tsg_id=tsg.id)

    # --- GET Request: Display the Main Form ---
    else:
        grades_queryset = Grade.objects.filter(teacher_assignment=tsg).select_related('student__user')
        grades_dict = {g.student.id: g for g in grades_queryset}
        attendance_history = Attendance.objects.filter(
            teacher_assignment=tsg
        ).order_by('-date').select_related('student__user')
        materials = Material.objects.filter(
            teacher_assignment=tsg
        ).order_by('-uploaded_at')
        
        sg_columns = tsg.get_sg_field_names()
        
        # Подготавливаем контекст с информацией о лимитах для шаблона
        grade_settings = GlobalGradeSettings.load()
        context = {
            'tsg': tsg,
            'students': students,
            'grades': grades_dict,
            'attendance_history': attendance_history,
            'materials': materials,
            'sg_columns': sg_columns,
            'setting_num_sg': False,
            'midterm_limit': grade_settings.midterm_limit,
            'final_limit': grade_settings.final_limit,
            'midterm_editable': grade_settings.midterm_limit is not None,
            'final_editable': grade_settings.final_limit is not None,
        }
        return render(request, 'school/grade_form.html', context)



def is_admin(user):
    return user.is_superuser

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    query = request.GET.get('q', '')
    students = StudentProfile.objects.select_related('user').prefetch_related('groups')
    teachers = CustomUser.objects.filter(role='teacher')
    teacher_assignments = TeacherAssignment.objects.select_related('subject', 'group', 'teacher')
    groups = Group.objects.all()
    
    if query:
        students = students.filter(
            Q(user__username__icontains=query) |
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query)
        )
        teachers = teachers.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )
        groups = groups.filter(name__icontains=query) # Фильтрация групп по запросу

    return render(request, 'school/admin_dashboard.html', {
        'students': students,
        'teachers': teachers,
        'teacher_assignments': teacher_assignments,
        'groups': groups,
        'query': query
    })

@login_required
@user_passes_test(is_admin)
def student_detail(request, student_id):
    student = get_object_or_404(
        StudentProfile.objects.select_related('user'),
        id=student_id
    )
    
    grades = Grade.objects.filter(student=student).select_related(
        'teacher_assignment__teacher',
        'teacher_assignment__subject',
        'teacher_assignment__group'
    )
    
    attendance_records = Attendance.objects.filter(student=student).select_related(
        'teacher_assignment__teacher',
        'teacher_assignment__subject'
    ).order_by('date')
    
    missed_lessons_by_subject = defaultdict(lambda: {'dates': [], 'total': 0})
    for att in attendance_records:
        subject_name = att.teacher_assignment.subject.name
        missed_lessons_by_subject[subject_name]['dates'].append(att.date)
        missed_lessons_by_subject[subject_name]['total'] += att.missed_lessons
    
    # Подготовка данных оценок для админки, аналогично student_subject_detail
    grades_data = []
    for grade in grades:
         sg_scores = grade.get_sg_scores()
         grades_data.append({
            'grade_obj': grade, # Передаем весь объект для доступа к другим полям
            'activity': grade.activity,
            'sg_scores': sg_scores,
            'sg_total': grade.calculate_sg_total(),
            'midterm': grade.midterm,
            'final': grade.final,
            'total': grade.total
         })

    return render(request, 'school/student_detail.html', {
        'student': student,
        'grades_data': grades_data, # Передаем подготовленные данные
        'missed_lessons_by_subject': dict(missed_lessons_by_subject),
    })

@login_required
@user_passes_test(is_admin)
def teacher_detail(request, teacher_id):
    teacher = get_object_or_404(CustomUser, id=teacher_id, role='teacher')
    assignments = TeacherAssignment.objects.filter(teacher=teacher).select_related('subject', 'group')
    subjects = set(assignment.subject for assignment in assignments)
    groups = set(assignment.group for assignment in assignments)
    
    return render(request, 'school/teacher_detail.html', {
        'teacher': teacher,
        'assignments': assignments,
        'subjects': subjects,
        'groups': groups
    })

@login_required
@user_passes_test(is_admin)
def group_detail(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    # Используем related_name 'groups' из StudentProfile
    students = group.students.all().select_related('user') 
    assignments = TeacherAssignment.objects.filter(group=group).select_related('subject', 'teacher')
    
    return render(request, 'school/group_detail.html', {
        'group': group,
        'students': students,
        'assignments': assignments
    })






# views.py
import pandas as pd
import tempfile
import sys
import os
import time
from io import StringIO
from django.http import JsonResponse
from django.views.generic import TemplateView
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.core.management import call_command, CommandError
from school.models import Group
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError

@method_decorator(staff_member_required, name='dispatch')
class ImportExcelView(TemplateView):
    template_name = 'admin/import_excel.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['groups'] = Group.objects.all()
        return context
    
    def get(self, request, *args, **kwargs):
        """Обработка GET запросов"""
        return super().get(request, *args, **kwargs)
    



















    def post(self, request, *args, **kwargs):
        try:
            excel_file = request.FILES.get('excel_file')
            group_name = request.POST.get('group_name', '').strip()
            preview_only = request.POST.get('preview_only') == '1'
            overwrite = request.POST.get('overwrite_existing') == 'on'
        
            if not excel_file:
                return JsonResponse({'success': False, 'error': 'Пожалуйста, выберите файл Excel.'})
        
            if not group_name:
                return JsonResponse({'success': False, 'error': 'Пожалуйста, введите название группы.'})
        
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                for chunk in excel_file.chunks():
                   tmp_file.write(chunk)
                tmp_file_path = tmp_file.name
    
            try:
                if preview_only:
                    preview_data = self.preview_excel_file(tmp_file_path, group_name)
                    return JsonResponse({
                        'success': True,
                        'preview': preview_data,
                        'message': 'Предварительный просмотр файла'
                    })
                else:
                    result = self.execute_import_command_sync(tmp_file_path, group_name, overwrite)
                    return JsonResponse(result)
            finally:
                try:
                    os.unlink(tmp_file_path)
                except OSError:
                    pass
                
        except Exception as e:
            import traceback
            error_details = f"{str(e)}\n{traceback.format_exc()}"
            logger.error(f"Ошибка в post view: {error_details}")
            return JsonResponse({
                'success': False,
                'error': f'Ошибка на сервере: {str(e)}',
                'details': error_details[:1000]
            })
    























    def preview_excel_file(self, file_path, group_name):
        try:
            df = pd.read_excel(file_path, header=None)
            
            preview = {
                'file_info': {
                    'total_rows': len(df),
                    'total_columns': len(df.columns) if len(df) > 0 else 0,
                },
                'group_name': group_name,
                'analysis': {
                    'student_column': 'A (Имя студента)',
                    'email_column': 'B (Email)',
                    'potential_grade_columns_count': max(0, len(df.columns) - 2),
                    'group_exists': Group.objects.filter(name=group_name).exists(),
                    'potential_issues': []
                },
                'students_preview': [],
                'grade_columns_sample': [],
                'estimated_time': self._estimate_import_time(df)
            }
    
            if len(df) < 4:
                preview['analysis']['potential_issues'].append('Файл содержит менее 4 строк')
            if len(df.columns) < 3:
                preview['analysis']['potential_issues'].append('Файл содержит менее 3 колонок')
    
            if len(df) > 3:
                data_start_row = 3
                for i in range(data_start_row, min(len(df), data_start_row + 5)):
                    student_cell = df.iloc[i, 0] if len(df.columns) > 0 else None
                    if pd.notna(student_cell) and isinstance(student_cell, str):
                        student_name = str(student_cell).strip()
                        if student_name and student_name.lower() not in ['nan', 'name', '', 'фамилия имя отчество']:
                            preview['students_preview'].append(student_name)
    
                total_students = 0
                for i in range(data_start_row, len(df)):
                    student_cell = df.iloc[i, 0] if len(df.columns) > 0 else None
                    if pd.notna(student_cell) and isinstance(student_cell, str):
                        student_name = str(student_cell).strip()
                        if student_name and student_name.lower() not in ['nan', 'name', '', 'фамилия имя отчество']:
                            total_students += 1
                preview['analysis']['estimated_students'] = total_students
    
            if len(df) > 1:
                subject_row = df.iloc[1]
                subject_count = 0
                for col_idx in range(2, len(subject_row)):
                    subject_cell = subject_row.iloc[col_idx] if col_idx < len(subject_row) else None
                    if pd.notna(subject_cell):
                        subject_name = str(subject_cell).strip()
                        if subject_name and subject_name.lower() not in ['gpa', 'total', 'credit', 'credits']:
                            if subject_count < 10:
                                preview['grade_columns_sample'].append(subject_name)
                            subject_count += 1
                preview['analysis']['total_subjects'] = subject_count
    
            return preview
    
        except Exception as e:
            error_msg = f'Error parsing file: {str(e)}'
            logger.error(f"Ошибка preview_excel_file: {error_msg}")
            return {'error': error_msg}
    









    
    def _estimate_import_time(self, df):
        try:
            estimated_students = max(0, len(df) - 3) if len(df) > 3 else 0
            estimated_subjects = max(0, len(df.columns) - 2) if len(df.columns) > 2 else 0
            estimated_operations = estimated_students * estimated_subjects
            base_time_per_operation = 0.05
            overhead_time = 30
            estimated_seconds = (estimated_operations * base_time_per_operation) + overhead_time
            
            if estimated_seconds < 60:
                return f"менее минуты ({estimated_seconds:.0f} сек.)"
            elif estimated_seconds < 3600:
                return f"{estimated_seconds/60:.1f} минут"
            else:
                return f"{estimated_seconds/3600:.1f} часов"
        except Exception as e:
            logger.error(f"Ошибка _estimate_import_time: {e}")
            return "неизвестно"
                






        except Exception as e:
            print(f"[DEBUG] error _estimate_import_time: {e}")
            return "неизвестно"

    def execute_import_command_sync(self, file_path, group_name, overwrite):
        logger.debug(f"Начало execute_import_command_sync для файла {file_path}")
        
        def run_command():
            old_stdout = sys.stdout
            old_stderr = sys.stderr
            sys.stdout = captured_output = StringIO()
            sys.stderr = captured_errors = StringIO()
            
            try:
                logger.debug("Внутри run_command, запуск команды...")
                args = ['--file', file_path, '--group', group_name]
                if overwrite:
                    args.append('--overwrite')
                
                call_command('import_excel_data', *args)
                
                output = captured_output.getvalue()
                errors = captured_errors.getvalue()
                
                logger.debug(f"Команда завершена. Output length: {len(output)}")
                
                if "Traceback" in output or "Error" in output or "Exception" in output:
                    return {
                        'success': False,
                        'error': 'Ошибка во время импорта',
                        'details': output[:2000] + "..." if len(output) > 2000 else output
                    }
                elif errors and errors.strip():
                    return {
                        'success': False,
                        'error': 'Ошибка во время импорта (stderr)',
                        'details': errors[:2000] + "..." if len(errors) > 2000 else errors
                    }
                else:
                    return {
                        'success': True,
                        'message': 'Импорт успешно завершен',
                        'details': output[:2000] + "..." if len(output) > 2000 else output
                    }
    
            except CommandError as e:
                output = captured_output.getvalue()
                errors = captured_errors.getvalue()
                logger.error(f"CommandError в run_command: {e}")
                return {
                    'success': False,
                    'error': f'Ошибка команды импорта: {str(e)}',
                    'details': (output[:1000] + "..." if len(output) > 1000 else output) +
                               ("\nErrors: " + (errors[:1000] + "..." if len(errors) > 1000 else errors) if errors.strip() else "")
                }
            except Exception as e:
                import traceback
                output = captured_output.getvalue()
                errors = captured_errors.getvalue()
                tb = traceback.format_exc()
                logger.error(f"Exception в run_command: {e}", exc_info=True)
                return {
                    'success': False,
                    'error': f'Неожиданная ошибка во время импорта: {str(e)}',
                    'details': f"Output: {(output[:500] + '...' if len(output) > 500 else output)}\n"
                               f"Errors: {(errors[:500] + '...' if len(errors) > 500 else errors)}\n"
                               f"Traceback: {(tb[:1000] + '...' if len(tb) > 1000 else tb)}"
                }
            finally:
                sys.stdout = old_stdout
                sys.stderr = old_stderr
                logger.debug("Восстановлены стандартные потоки в run_command")
    
        try:
            logger.debug("Запуск ThreadPoolExecutor...")
            with ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(run_command)
                result = future.result(timeout=1800)
                logger.debug("Результат получен из ThreadPoolExecutor")
            return result
            
        except FutureTimeoutError:
            logger.warning("Таймаут в execute_import_command_sync")
            return {
                'success': False,
                'error': 'Импорт превысил время ожидания (30 минут). '
                         'Процесс может быть продолжен в фоне, но результат не будет доступен в интерфейсе. '
                         'Проверьте логи сервера Django.'
            }
        except Exception as e:
            logger.error(f"Ошибка в execute_import_command_sync: {e}")
            return {
                'success': False,
                'error': f'Ошибка при запуске импорта: {str(e)}'
            }




























@login_required
@user_passes_test(is_tutor)
def tutor_dashboard(request):
    """
    Главная страница тьютора.
    Отображает список групп, назначенных тьютору.
    """
    try:
        tutor_profile = request.user.tutor_profile
        # Получаем группы тьютора
        groups = tutor_profile.groups.all()
    except TutorProfile.DoesNotExist:
        # Если профиль не создан, показываем пустой список
        groups = Group.objects.none()
        messages.error(request, 'Профиль тьютора не настроен. Обратитесь к администратору.')

    context = {
        'groups': groups,
    }
    return render(request, 'tutor/dashboard.html', context)


@login_required
@user_passes_test(is_tutor)
def tutor_group_detail(request, group_id):
    """
    Страница деталей группы для тьютора.
    Отображает список студентов и предметы.
    """
    try:
        tutor_profile = request.user.tutor_profile
        # Убедимся, что тьютор имеет доступ к этой группе
        group = get_object_or_404(tutor_profile.groups, id=group_id)
    except TutorProfile.DoesNotExist:
        raise PermissionDenied("Профиль тьютора не настроен.")
    except Http404:
        raise PermissionDenied("У вас нет доступа к этой группе.")

    # Получаем студентов группы
    students = group.students.all().order_by('user__last_name', 'user__first_name')
    
    # Получаем уникальные предметы и преподавателей для этой группы
    # Используем Prefetch для оптимизации запросов
    teacher_assignments = TeacherAssignment.objects.filter(
        group=group
    ).select_related(
        'subject', 'teacher', 'semester'
    ).order_by('-semester__number', 'subject__name')

    # Группируем по семестрам, затем по предметам
    subjects_by_semester = {}
    for ta in teacher_assignments:
        semester = ta.semester
        subject_name = ta.subject.name
        teacher_name = ta.teacher.get_full_name()
        
        if semester not in subjects_by_semester:
            subjects_by_semester[semester] = {}
        
        if subject_name not in subjects_by_semester[semester]:
            subjects_by_semester[semester][subject_name] = []
            
        if teacher_name not in subjects_by_semester[semester][subject_name]:
            subjects_by_semester[semester][subject_name].append(teacher_name)

    # Получаем материалы для этой группы
    materials = Material.objects.filter(
        teacher_assignment__group=group
    ).select_related(
        'teacher_assignment__subject',
        'teacher_assignment__teacher',
        'teacher_assignment__semester'
    ).order_by('-uploaded_at')

    context = {
        'group': group,
        'students': students,
        'subjects_by_semester': subjects_by_semester,
        'materials': materials,
    }
    return render(request, 'tutor/group_detail.html', context)





@login_required
@user_passes_test(is_tutor)
def tutor_student_grades(request, student_id):
    """
    Страница оценок и посещаемости студента для тьютора.
    Только просмотр, редактирование запрещено.
    """
    try:
        tutor_profile = request.user.tutor_profile
        # Получаем студента и проверяем, что он в группе тьютора
        student_profile = get_object_or_404(
            StudentProfile.objects.prefetch_related('groups'),
            id=student_id
        )
        # Проверка доступа: студент должен быть в одной из групп тьютора
        tutor_groups = tutor_profile.groups.all()
        if not student_profile.groups.filter(id__in=tutor_groups).exists():
            raise PermissionDenied("У вас нет доступа к этому студенту.")

    except TutorProfile.DoesNotExist:
        raise PermissionDenied("Профиль тьютора не настроен.")
    except Http404:
        raise PermissionDenied("Студент не найден.")

    # Получаем оценки студента
    grades = Grade.objects.filter(student=student_profile).select_related(
        'teacher_assignment__subject', 'teacher_assignment__teacher', 'teacher_assignment__semester'
    )

    # Получаем посещаемость студента
    attendance_records = Attendance.objects.filter(student=student_profile).select_related(
        'teacher_assignment__subject', 'teacher_assignment__teacher', 'teacher_assignment__semester'
    ).order_by('-date')

    # Создаем сводку по пропускам: {teacher_assignment_id: total_missed_lessons}
    attendance_summary = {}
    for attendance in attendance_records:
        ta_id = attendance.teacher_assignment.id
        if ta_id not in attendance_summary:
            attendance_summary[ta_id] = 0
        attendance_summary[ta_id] += attendance.missed_lessons

    context = {
        'student': student_profile,
        'grades': grades,
        'attendance_records': attendance_records,
        'attendance_summary': attendance_summary,
    }
    return render(request, 'tutor/student_grades.html', context)







def is_scheduler(user):
    return user.is_authenticated and user.role == CustomUser.SCHEDULER



# Исправленный schedule_list (использует weekday, week_type)
@login_required
@user_passes_test(is_scheduler)
def schedule_list(request):
    """
    Список всех записей в расписании для текущего Scheduler'а.
    """
    # Получаем QuerySet без фильтрации по умолчанию
    entries = ScheduleEntry.objects.filter(scheduler=request.user).select_related(
        'group', 'teacher', 'subject', 'room'
    ).order_by('week_type', 'weekday', 'time_slot')

    # Фильтрация (опционально)
    week_filter = request.GET.get('week')
    group_filter = request.GET.get('group')
    teacher_filter = request.GET.get('teacher')
    subject_filter = request.GET.get('subject')

    if week_filter:
        entries = entries.filter(week_type=week_filter)
    if group_filter:
        entries = entries.filter(group_id=group_filter)
    if teacher_filter:
        entries = entries.filter(teacher_id=teacher_filter)
    if subject_filter:
        entries = entries.filter(subject_id=subject_filter)

    groups = Group.objects.all()
    teachers = CustomUser.objects.filter(role='teacher')
    subjects = Subject.objects.all()
    week_choices = ScheduleEntry.WEEK_CHOICES
    time_slots = ScheduleEntry.TIME_SLOT_CHOICES

    context = {
        'entries': entries,
        'groups': groups,
        'teachers': teachers,
        'subjects': subjects,
        'week_choices': week_choices,
        'time_slots': time_slots,
        'week_filter': week_filter,
        'group_filter': group_filter,
        'teacher_filter': teacher_filter,
        'subject_filter': subject_filter,
    }
    return render(request, 'scheduler/schedule_list.html', context)





@login_required
@user_passes_test(is_scheduler)
def schedule_edit(request):
    """
    Редактирование всего существующего расписания с предзаполнением данных.
    """
    scheduler = request.user
    
    # Получаем все существующие записи для этого расписателя
    existing_entries = ScheduleEntry.objects.filter(scheduler=scheduler).select_related(
        'group', 'teacher', 'subject', 'room'
    ).order_by('week_type', 'weekday', 'time_slot')
    
    if request.method == 'POST':
        logger.info(f"POST request for schedule editing")
        logger.info(f"POST data: {dict(request.POST)}")
        
        try:
            with transaction.atomic():
                # Удаляем ВСЕ старые записи этого scheduler'а
                deleted_count = ScheduleEntry.objects.filter(scheduler=scheduler).delete()[0]
                logger.info(f"Deleted {deleted_count} old entries")

                saved_count = 0
                errors = []

                # Обрабатываем все ключи POST
                teacher_keys = [key for key in request.POST.keys() if key.startswith('teacher_')]
                
                for key in teacher_keys:
                    teacher_id = request.POST.get(key)
                    if not teacher_id:
                        continue
                    
                    try:
                        # Парсим ключ: teacher_top_0_1_1 (week_type_weekday_time_slot_entry_num)
                        parts = key.split('_')
                        if len(parts) != 5:
                            logger.warning(f"Неправильный формат ключа: {key}")
                            continue
                            
                        week_type = parts[1]  # 'top' или 'bottom'
                        weekday = int(parts[2])
                        time_slot = int(parts[3])
                        entry_index = int(parts[4])  # номер записи
                        
                        # Получаем соответствующие group, subject и room
                        group_key = f"group_{week_type}_{weekday}_{time_slot}_{entry_index}"
                        subject_key = f"subject_{week_type}_{weekday}_{time_slot}_{entry_index}"
                        room_key = f"room_{week_type}_{weekday}_{time_slot}_{entry_index}"
                        
                        group_id = request.POST.get(group_key)
                        subject_id = request.POST.get(subject_key)
                        room_id = request.POST.get(room_key)

                        logger.info(f"Processing entry {entry_index}: {key} -> group={group_id}, teacher={teacher_id}, subject={subject_id}, room={room_id}")

                        # Проверяем, что все поля заполнены
                        if group_id and teacher_id and subject_id and room_id:
                            try:
                                # Проверяем существование объектов
                                group = Group.objects.get(id=group_id)
                                teacher = CustomUser.objects.get(id=teacher_id, role=CustomUser.TEACHER)
                                subject = Subject.objects.get(id=subject_id)
                                room = Room.objects.get(id=room_id)
                                
                                # Создаем НОВУЮ запись (не обновляем старую)
                                entry = ScheduleEntry.objects.create(
                                    scheduler=scheduler,
                                    weekday=weekday,
                                    week_type=week_type,
                                    time_slot=time_slot,
                                    group=group,
                                    teacher=teacher,
                                    subject=subject,
                                    room=room
                                )
                                saved_count += 1
                                logger.info(f"Created new entry: {entry}")
                                
                            except (Group.DoesNotExist, CustomUser.DoesNotExist, Subject.DoesNotExist, Room.DoesNotExist) as e:
                                error_msg = f"Не найден объект для entry {entry_index} {key}: {e}"
                                errors.append(error_msg)
                                logger.error(error_msg)
                                
                        elif group_id or teacher_id or subject_id or room_id:
                            logger.warning(f"Неполные данные для entry {entry_index} {key}")
                            
                    except (ValueError, IndexError) as e:
                        error_msg = f"Ошибка обработки {key}: {e}"
                        errors.append(error_msg)
                        logger.error(error_msg)
                    except Exception as e:
                        error_msg = f"Неожиданная ошибка для {key}: {e}"
                        errors.append(error_msg)
                        logger.error(error_msg, exc_info=True)

                if errors:
                    messages.warning(request, f'Сохранено {saved_count} записей. Ошибки: {"; ".join(errors[:5])}...')
                else:
                    messages.success(request, f'Расписание успешно обновлено! Сохранено {saved_count} записей.')
                
                logger.info(f"Successfully processed {saved_count} entries")
                return redirect('schedule_list')

        except Exception as e:
            error_msg = f'Ошибка при обновлении расписания: {e}'
            logger.error(error_msg, exc_info=True)
            messages.error(request, error_msg)

    # GET запрос - показываем форму с предзаполненными данными
    groups = Group.objects.all()
    teachers = CustomUser.objects.filter(role=CustomUser.TEACHER)
    subjects = Subject.objects.all()
    rooms = Room.objects.all()
    time_slots = ScheduleEntry.TIME_SLOT_CHOICES

    # ИСПРАВЛЕННАЯ структура данных для предзаполнения формы
    schedule_data = {}
    
    # Группируем записи по ключу week_type_weekday_time_slot
    for entry in existing_entries:
        key = f"{entry.week_type}_{entry.weekday}_{entry.time_slot}"
        if key not in schedule_data:
            schedule_data[key] = []
        schedule_data[key].append({
            'group_id': entry.group.id if entry.group else None,
            'subject_id': entry.subject.id if entry.subject else None,
            'teacher_id': entry.teacher.id if entry.teacher else None,
            'room_id': entry.room.id if entry.room else None,
        })

    # Логируем структуру данных для отладки
    logger.info(f"Schedule data structure: {list(schedule_data.keys())}")
    logger.info(f"Existing entries count: {existing_entries.count()}")
    
    # Выводим детальную информацию для отладки
    for key, entries in schedule_data.items():
        logger.info(f"Key {key}: {len(entries)} entries")

    context = {
        'groups': groups,
        'teachers': teachers,
        'subjects': subjects,
        'rooms': rooms,
        'time_slots': time_slots,
        'schedule_data': schedule_data,
        'existing_entries': existing_entries,
    }
    logger.debug(f"Context: groups={groups.count()}, subjects={subjects.count()}, schedule_data_keys={list(schedule_data.keys())}")


    return render(request, 'scheduler/schedule_edit.html', context)





@login_required
@user_passes_test(is_scheduler)
def schedule_update(request, entry_id):
    """
    Редактирование записи в расписании.
    """
    entry = get_object_or_404(ScheduleEntry, id=entry_id, scheduler=request.user)

    if request.method == 'POST':
        form = ScheduleEntryForm(request.POST, instance=entry)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Запись в расписании успешно обновлена.')
                return redirect('schedule_list')
            except forms.ValidationError as e: # Убедитесь, что 'forms' импортирован
                 messages.error(request, f"Ошибка валидации: {e}")
            except Exception as e:
                 messages.error(request, f"Ошибка при сохранении: {e}")
        else:
             messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
    else:
        form = ScheduleEntryForm(instance=entry)

    context = {
        'form': form,
        'action': 'Обновить',
        'entry': entry,
    }
    return render(request, 'scheduler/schedule_form.html', context)






@login_required
@user_passes_test(is_scheduler)
def schedule_group(request, group_id):
    """
    Страница редактирования/создания расписания для конкретной группы.
    """
    group = get_object_or_404(Group, id=group_id)
    scheduler = request.user

    # Получаем все возможные дни недели и временные слоты
    weekdays = ScheduleEntry.WEEKDAY_CHOICES
    time_slots = ScheduleEntry.TIME_SLOT_CHOICES

    # Получаем существующие записи для этой группы и расписателя
    existing_entries = ScheduleEntry.objects.filter(
        group=group,
        scheduler=scheduler
    ).select_related('teacher', 'subject', 'room')

    # Создаем структуру данных для шаблона
    schedule_data = {}
    for entry in existing_entries:
        key = f"{entry.week_type}_{entry.weekday}_{entry.time_slot}"
        schedule_data[key] = {
            'teacher_id': entry.teacher.id,
            'subject_id': entry.subject.id,
            'room_id': entry.room.id,
        }

    if request.method == 'POST':
        logger.info(f"POST request for group {group_id}")
        logger.info(f"POST data: {dict(request.POST)}")
        
        try:
            with transaction.atomic():
                # Удаляем старые записи
                deleted_count = ScheduleEntry.objects.filter(
                    group=group, 
                    scheduler=scheduler
                ).delete()[0]
                logger.info(f"Deleted {deleted_count} old entries")

                saved_count = 0
                errors = []

                # Обрабатываем данные формы
                processed_keys = set()
                
                for key in request.POST.keys():
                    if key.startswith('teacher_') and key not in processed_keys:
                        processed_keys.add(key)
                        
                        teacher_id = request.POST.get(key)
                        if not teacher_id:
                            continue
                        
                        try:
                            # Парсим ключ: teacher_top_0_1
                            parts = key.split('_')
                            if len(parts) != 4:
                                logger.warning(f"Неправильный формат ключа: {key}")
                                continue
                                
                            week_type = parts[1]  # 'top' или 'bottom'
                            weekday = int(parts[2])
                            time_slot = int(parts[3])
                            
                            # Получаем соответствующие subject и room
                            subject_key = f"subject_{week_type}_{weekday}_{time_slot}"
                            room_key = f"room_{week_type}_{weekday}_{time_slot}"
                            
                            subject_id = request.POST.get(subject_key)
                            room_id = request.POST.get(room_key)

                            logger.info(f"Processing: {key} -> teacher={teacher_id}, subject={subject_id}, room={room_id}")

                            if teacher_id and subject_id and room_id:
                                try:
                                    # Проверяем существование объектов
                                    teacher = CustomUser.objects.get(id=teacher_id, role=CustomUser.TEACHER)
                                    subject = Subject.objects.get(id=subject_id)
                                    room = Room.objects.get(id=room_id)
                                    
                                    # Создаем запись
                                    entry = ScheduleEntry.objects.create(
                                        scheduler=scheduler,
                                        weekday=weekday,
                                        week_type=week_type,
                                        time_slot=time_slot,
                                        group=group,
                                        teacher=teacher,
                                        subject=subject,
                                        room=room
                                    )
                                    saved_count += 1
                                    logger.info(f"Created entry: {entry}")
                                    
                                except (CustomUser.DoesNotExist, Subject.DoesNotExist, Room.DoesNotExist) as e:
                                    error_msg = f"Не найден объект для {key}: {e}"
                                    errors.append(error_msg)
                                    logger.error(error_msg)
                                    
                            else:
                                logger.warning(f"Неполные данные для {key}: teacher={teacher_id}, subject={subject_id}, room={room_id}")
                                
                        except (ValueError, IndexError) as e:
                            error_msg = f"Ошибка обработки {key}: {e}"
                            errors.append(error_msg)
                            logger.error(error_msg)
                        except Exception as e:
                            error_msg = f"Неожиданная ошибка для {key}: {e}"
                            errors.append(error_msg)
                            logger.error(error_msg, exc_info=True)

                if errors:
                    messages.warning(request, f'Сохранено {saved_count} записей. Ошибки: {"; ".join(errors[:3])}...')
                else:
                    messages.success(request, f'Расписание успешно сохранено! Создано {saved_count} записей.')
                
                logger.info(f"Successfully saved {saved_count} entries")
                return redirect('schedule_group', group_id=group.id)

        except Exception as e:
            error_msg = f'Ошибка при сохранении расписания: {e}'
            logger.error(error_msg, exc_info=True)
            messages.error(request, error_msg)

    # GET запрос - показываем форму
    teachers = CustomUser.objects.filter(role=CustomUser.TEACHER)
    subjects = Subject.objects.all()
    rooms = Room.objects.all()

    context = {
        'group': group,
        'weekdays': weekdays,
        'time_slots': time_slots,
        'schedule_data': schedule_data,
        'teachers': teachers,
        'subjects': subjects,
        'rooms': rooms,
    }
    
    return render(request, 'scheduler/schedule_group.html', context)




@login_required
@user_passes_test(is_scheduler)
def scheduler_dashboard(request):
    """
    Главная страница Scheduler'а.
    """
    if request.method == 'POST':
        form = GroupScheduleForm(request.POST)
        if form.is_valid():
            group = form.cleaned_data['group']
            # Перенаправляем на страницу редактирования расписания для группы
            return redirect('schedule_group', group_id=group.id)
    else:
        form = GroupScheduleForm()

    context = {
        'form': form,
    }
    return render(request, 'scheduler/dashboard.html', context)




@login_required
@user_passes_test(is_scheduler)
def schedule_create(request):
    """
    Создание нового расписания - показываем полную таблицу для обеих недель.
    """
    scheduler = request.user
    
    if request.method == 'POST':
        logger.info(f"POST request for schedule creation")
        logger.info(f"POST  {dict(request.POST)}")
        
        try:
            with transaction.atomic():
                # Удаляем старые записи для этого расписателя
                deleted_count = ScheduleEntry.objects.filter(scheduler=scheduler).delete()[0]
                logger.info(f"Deleted {deleted_count} old entries")

                saved_count = 0
                errors = []

                # Обрабатываем все ключи POST
                teacher_keys = [key for key in request.POST.keys() if key.startswith('teacher_')]
                
                for key in teacher_keys:
                    teacher_id = request.POST.get(key)
                    if not teacher_id:
                        continue
                    
                    try:
                        # Парсим ключ: teacher_top_0_1_1 (week_type_weekday_time_slot_entry_num)
                        parts = key.split('_')
                        if len(parts) != 5:
                            logger.warning(f"Неправильный формат ключа: {key}")
                            continue
                            
                        week_type = parts[1]  # 'top' или 'bottom'
                        weekday = int(parts[2])
                        time_slot = int(parts[3])
                        entry_index = int(parts[4])  # номер записи
                        
                        # Получаем соответствующие group, subject и room
                        group_key = f"group_{week_type}_{weekday}_{time_slot}_{entry_index}"
                        subject_key = f"subject_{week_type}_{weekday}_{time_slot}_{entry_index}"
                        room_key = f"room_{week_type}_{weekday}_{time_slot}_{entry_index}"
                        
                        group_id = request.POST.get(group_key)
                        subject_id = request.POST.get(subject_key)
                        room_id = request.POST.get(room_key)

                        logger.info(f"Processing entry {entry_index}: {key} -> group={group_id}, teacher={teacher_id}, subject={subject_id}, room={room_id}")

                        # Проверяем, что все поля заполнены
                        if group_id and teacher_id and subject_id and room_id:
                            try:
                                # Проверяем существование объектов
                                group = Group.objects.get(id=group_id)
                                teacher = CustomUser.objects.get(id=teacher_id, role=CustomUser.TEACHER)
                                subject = Subject.objects.get(id=subject_id)
                                room = Room.objects.get(id=room_id)
                                
                                # Создаем запись
                                entry = ScheduleEntry.objects.create(
                                    scheduler=scheduler,
                                    weekday=weekday,
                                    week_type=week_type,
                                    time_slot=time_slot,
                                    group=group,
                                    teacher=teacher,
                                    subject=subject,
                                    room=room
                                )
                                saved_count += 1
                                logger.info(f"Created entry: {entry}")
                                
                            except (Group.DoesNotExist, CustomUser.DoesNotExist, Subject.DoesNotExist, Room.DoesNotExist) as e:
                                error_msg = f"Не найден объект для entry {entry_index} {key}: {e}"
                                errors.append(error_msg)
                                logger.error(error_msg)
                                
                        elif group_id or teacher_id or subject_id or room_id:
                            logger.warning(f"Неполные данные для entry {entry_index} {key}")
                            
                    except (ValueError, IndexError) as e:
                        error_msg = f"Ошибка обработки {key}: {e}"
                        errors.append(error_msg)
                        logger.error(error_msg)
                    except Exception as e:
                        error_msg = f"Неожиданная ошибка для {key}: {e}"
                        errors.append(error_msg)
                        logger.error(error_msg, exc_info=True)

                if errors:
                    messages.warning(request, f'Сохранено {saved_count} записей. Ошибки: {"; ".join(errors[:5])}...')
                else:
                    messages.success(request, f'Расписание успешно сохранено! Создано {saved_count} записей.')
                
                logger.info(f"Successfully saved {saved_count} entries")
                return redirect('schedule_list')

        except Exception as e:
            error_msg = f'Ошибка при сохранении расписания: {e}'
            logger.error(error_msg, exc_info=True)
            messages.error(request, error_msg)

    # GET запрос - показываем форму
    groups = Group.objects.all()
    teachers = CustomUser.objects.filter(role=CustomUser.TEACHER)
    subjects = Subject.objects.all()
    rooms = Room.objects.all()
    time_slots = ScheduleEntry.TIME_SLOT_CHOICES

    context = {
        'groups': groups,
        'teachers': teachers,
        'subjects': subjects,
        'rooms': rooms,
        'time_slots': time_slots,
    }
    return render(request, 'scheduler/schedule_create.html', context)














@login_required
@user_passes_test(is_scheduler)
def export_schedule_teacher(request):
    teacher_id = request.GET.get("teacher")
    if not teacher_id:
        messages.error(request, "Выберите преподавателя для экспорта.")
        return redirect("export_teacher_form")

    teacher = get_object_or_404(CustomUser, id=teacher_id, role="teacher")
    entries = ScheduleEntry.objects.filter(teacher=teacher).select_related(
        "group", "subject", "room"
    ).order_by("week_type", "weekday", "time_slot")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{teacher.get_full_name()}"

    row = 1
    ws.cell(row=row, column=1, value=f"Преподаватель: {teacher.get_full_name()}")
    row += 2

    # Правильный порядок дней недели
    weekdays = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
    
    # Группируем по неделям в правильном порядке
    weeks = ["top", "bottom"]  # Сначала верхняя, потом нижняя
    
    for week_type in weeks:
        week_label = "Верхняя" if week_type == "top" else "Нижняя"
        week_entries = entries.filter(week_type=week_type)
        
        if not week_entries.exists():
            continue
            
        # Добавляем заголовок недели
        row += 1
        ws.cell(row=row, column=1, value=f"Неделя: {week_label}")
        row += 1
        
        # Проходим по дням недели в правильном порядке (0-4, понедельник-пятница)
        for weekday_num in range(5):  # 0, 1, 2, 3, 4
            day_entries = week_entries.filter(weekday=weekday_num)
            
            if not day_entries.exists():
                continue
                
            day_label = weekdays[weekday_num]
            
            # Добавляем заголовок дня
            row += 1
            ws.cell(row=row, column=1, value=day_label)
            row += 1
            
            # Добавляем заголовки таблицы
            headers = ["Время", "Предмет", "Группа", "Кабинет"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=row, column=col, value=header)
            row += 1
            
            # Добавляем записи для этого дня, отсортированные по времени
            for entry in day_entries.order_by('time_slot'):
                start, end = get_time_range_from_slot(entry.time_slot)
                values = [
                    f"{start}-{end}", 
                    entry.subject.name, 
                    str(entry.group), 
                    str(entry.room)
                ]
                for col, val in enumerate(values, start=1):
                    ws.cell(row=row, column=col, value=val)
                row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=schedule_teacher_{teacher.last_name}.xlsx'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_scheduler)
def export_schedule_student(request):
    group_id = request.GET.get("group")
    group = get_object_or_404(Group, pk=group_id)
    entries = ScheduleEntry.objects.filter(group=group).select_related(
        "teacher", "subject", "room"
    ).order_by("week_type", "weekday", "time_slot")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Группа {group.name}"

    row = 1
    ws.cell(row=row, column=1, value=f"Группа: {group.name}")
    row += 2

    # Правильный порядок дней недели (понедельник = 0, пятница = 4)
    weekdays = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
    
    # Группируем по неделям в правильном порядке
    weeks = ["top", "bottom"]  # Сначала верхняя, потом нижняя
    
    for week_type in weeks:
        week_label = "Верхняя" if week_type == "top" else "Нижняя"
        week_entries = entries.filter(week_type=week_type)
        
        if not week_entries.exists():
            continue
            
        # Добавляем заголовок недели
        row += 1
        ws.cell(row=row, column=1, value=f"Неделя: {week_label}")
        row += 1
        
        # Проходим по дням недели в правильном порядке (0-4, понедельник-пятница)
        for weekday_num in range(5):  # 0, 1, 2, 3, 4
            day_entries = week_entries.filter(weekday=weekday_num)
            
            if not day_entries.exists():
                continue
                
            day_label = weekdays[weekday_num]
            
            # Добавляем заголовок дня
            row += 1
            ws.cell(row=row, column=1, value=day_label)
            row += 1
            
            # Добавляем заголовки таблицы
            headers = ["Время", "Предмет", "Преподаватель", "Кабинет"]
            for col, header in enumerate(headers, start=1):
                ws.cell(row=row, column=col, value=header)
            row += 1
            
            # Добавляем записи для этого дня, отсортированные по времени
            for entry in day_entries.order_by('time_slot'):
                start, end = get_time_range_from_slot(entry.time_slot)
                values = [
                    f"{start}-{end}", 
                    entry.subject.name, 
                    entry.teacher.get_full_name(), 
                    str(entry.room)
                ]
                for col, val in enumerate(values, start=1):
                    ws.cell(row=row, column=col, value=val)
                row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=student_{group.name}_schedule.xlsx'
    wb.save(response)
    return response




@login_required
@user_passes_test(is_scheduler)
def schedule_delete(request, entry_id):
    """
    Удаление записи из расписания.
    """
    # Убедитесь, что запись принадлежит текущему расписателю
    entry = get_object_or_404(ScheduleEntry, id=entry_id, scheduler=request.user)

    if request.method == 'POST':
        entry.delete()
        messages.success(request, 'Запись в расписании успешно удалена.')
        return redirect('schedule_list') # Или другой подходящий URL

    # Для GET запроса показываем страницу подтверждения
    context = {
        'entry': entry,
    }
    return render(request, 'scheduler/schedule_confirm_delete.html', context)



@login_required
@user_passes_test(is_scheduler)
def export_student_form(request):
    groups = Group.objects.all()
    return render(request, 'scheduler/export_student.html', {
        'groups': groups,
        'weekdays': WEEKDAYS,
    })


@login_required
@user_passes_test(is_scheduler)
def export_teacher_form(request):
    teachers = CustomUser.objects.filter(role='teacher')  # или CustomUser.TEACHER
    return render(request, 'scheduler/export_teacher.html', {
        'teachers': teachers,
        'weekdays': WEEKDAYS,
    })



def analyze_excel_file(file_path):
    """Анализирует Excel файл для предварительного просмотра"""
    try:
        df = pd.read_excel(file_path)
        preview_data = {
            'total_rows': len(df),
            'columns': list(df.columns),
            'sample_students': [],
            'column_analysis': {}
        }
        
        # Анализируем колонки
        for i, col in enumerate(df.columns):
            col_name = str(col).lower().strip()
            if i == 0:
                preview_data['column_analysis']['student_names'] = {
                    'column': col,
                    'sample_values': df[col].head(5).tolist()
                }
            elif any(keyword in col_name for keyword in ['midterm', 'activity', 'final', 'total', 'sg']):
                preview_data['column_analysis'][col_name] = {
                    'column': col,
                    'sample_values': df[col].head(5).tolist()
                }
        
        # Показываем первых 10 студентов
        first_col = df.columns[0]
        student_names = df[first_col].head(10).tolist()
        preview_data['sample_students'] = [str(name) for name in student_names if pd.notna(name)]
        
        return preview_data
    except Exception as e:
        return {'error': f'Ошибка при чтении файла: {str(e)}'}



@login_required
@user_passes_test(is_scheduler)
def schedule_group_enhanced(request, group_id):
    """
    Улучшенная версия страницы редактирования расписания для конкретной группы
    с проверкой конфликтов и визуальными индикаторами занятости.
    """
    group = get_object_or_404(Group, id=group_id)
    scheduler = request.user

    # Получаем все возможные дни недели и временные слоты
    weekdays = ScheduleEntry.WEEKDAY_CHOICES
    time_slots = ScheduleEntry.TIME_SLOT_CHOICES

    # Получаем существующие записи для этой группы и расписателя
    existing_entries = ScheduleEntry.objects.filter(
        group=group,
        scheduler=scheduler
    ).select_related('teacher', 'subject', 'room')

    # Создаем структуру данных для шаблона
    schedule_data = {}
    for entry in existing_entries:
        key = f"{entry.week_type}_{entry.weekday}_{entry.time_slot}"
        schedule_data[key] = {
            'teacher_id': entry.teacher.id,
            'subject_id': entry.subject.id,
            'room_id': entry.room.id,
        }

    # Получаем данные о занятости кабинетов и преподавателей
    occupied_rooms_data, occupied_teachers_data = get_occupancy_data(group_id)

    if request.method == 'POST':
        logger.info(f"POST request for group {group_id}")
        
        try:
            with transaction.atomic():
                # Удаляем старые записи
                deleted_count = ScheduleEntry.objects.filter(
                    group=group, 
                    scheduler=scheduler
                ).delete()[0]
                logger.info(f"Deleted {deleted_count} old entries")

                saved_count = 0
                errors = []
                conflicts = []

                # Проверяем конфликты перед сохранением
                for key in request.POST.keys():
                    if key.startswith('teacher_') and request.POST.get(key):
                        try:
                            parts = key.split('_')
                            if len(parts) != 4:
                                continue
                                
                            week_type = parts[1]
                            weekday = int(parts[2])
                            time_slot = int(parts[3])
                            
                            teacher_id = request.POST.get(key)
                            room_key = f"room_{week_type}_{weekday}_{time_slot}"
                            room_id = request.POST.get(room_key)

                            # Проверяем конфликты
                            room_conflict = check_room_conflict(
                                week_type, weekday, time_slot, room_id, group_id
                            )
                            teacher_conflict = check_teacher_conflict(
                                week_type, weekday, time_slot, teacher_id, group_id
                            )

                            if room_conflict:
                                conflicts.append(room_conflict)
                            if teacher_conflict:
                                conflicts.append(teacher_conflict)

                        except (ValueError, IndexError):
                            continue

                # Если есть критичные конфликты, не сохраняем
                if conflicts and not request.POST.get('force_save'):
                    messages.warning(
                        request, 
                        f'Обнаружены конфликты расписания: {"; ".join(conflicts[:3])}... '
                        'Используйте кнопку "Сохранить принудительно" для игнорирования конфликтов.'
                    )
                    # Возвращаем форму с текущими данными и информацией о конфликтах
                    context = prepare_schedule_context(
                        group, weekdays, time_slots, schedule_data, 
                        occupied_rooms_data, occupied_teachers_data, conflicts
                    )
                    return render(request, 'scheduler/schedule_group_enhanced.html', context)

                # Обрабатываем и сохраняем данные формы
                processed_keys = set()
                
                for key in request.POST.keys():
                    if key.startswith('teacher_') and key not in processed_keys:
                        processed_keys.add(key)
                        
                        teacher_id = request.POST.get(key)
                        if not teacher_id:
                            continue
                        
                        try:
                            parts = key.split('_')
                            if len(parts) != 4:
                                continue
                                
                            week_type = parts[1]
                            weekday = int(parts[2])
                            time_slot = int(parts[3])
                            
                            subject_key = f"subject_{week_type}_{weekday}_{time_slot}"
                            room_key = f"room_{week_type}_{weekday}_{time_slot}"
                            
                            subject_id = request.POST.get(subject_key)
                            room_id = request.POST.get(room_key)

                            if teacher_id and subject_id and room_id:
                                try:
                                    teacher = CustomUser.objects.get(id=teacher_id, role=CustomUser.TEACHER)
                                    subject = Subject.objects.get(id=subject_id)
                                    room = Room.objects.get(id=room_id)
                                    
                                    entry = ScheduleEntry.objects.create(
                                        scheduler=scheduler,
                                        weekday=weekday,
                                        week_type=week_type,
                                        time_slot=time_slot,
                                        group=group,
                                        teacher=teacher,
                                        subject=subject,
                                        room=room
                                    )
                                    saved_count += 1
                                    
                                except (CustomUser.DoesNotExist, Subject.DoesNotExist, Room.DoesNotExist) as e:
                                    error_msg = f"Не найден объект для {key}: {e}"
                                    errors.append(error_msg)
                                    
                        except (ValueError, IndexError) as e:
                            errors.append(f"Ошибка обработки {key}: {e}")

                if errors:
                    messages.warning(request, f'Сохранено {saved_count} записей. Ошибки: {"; ".join(errors[:3])}')
                else:
                    messages.success(request, f'Расписание успешно сохранено! Создано {saved_count} записей.')
                
                return redirect('schedule_group_enhanced', group_id=group.id)

        except Exception as e:
            logger.error(f'Ошибка при сохранении расписания: {e}', exc_info=True)
            messages.error(request, f'Ошибка при сохранении расписания: {e}')

    # GET запрос - показываем форму
    context = prepare_schedule_context(
        group, weekdays, time_slots, schedule_data, 
        occupied_rooms_data, occupied_teachers_data
    )
    
    return render(request, 'scheduler/schedule_group_enhanced.html', context)


def prepare_schedule_context(group, weekdays, time_slots, schedule_data, 
                           occupied_rooms_data, occupied_teachers_data, conflicts=None):
    """
    Подготавливает контекст для шаблона расписания.
    """
    teachers = CustomUser.objects.filter(role=CustomUser.TEACHER)
    subjects = Subject.objects.all()
    rooms = Room.objects.all()

    context = {
        'group': group,
        'weekdays': weekdays,
        'time_slots': time_slots,
        'schedule_data': schedule_data,
        'teachers': teachers,
        'subjects': subjects,
        'rooms': rooms,
        'occupied_rooms_data': json.dumps(occupied_rooms_data),
        'occupied_teachers_data': json.dumps(occupied_teachers_data),
    }
    
    if conflicts:
        context['conflicts'] = conflicts
    
    return context


def get_occupancy_data(current_group_id=None):
    """
    Получает данные о занятости кабинетов и преподавателей для всех групп,
    кроме текущей редактируемой группы.
    
    Возвращает:
        tuple: (occupied_rooms_data, occupied_teachers_data)
    """
    # Получаем все записи расписания, кроме текущей группы
    query = ScheduleEntry.objects.select_related('teacher', 'room', 'group')
    if current_group_id:
        query = query.exclude(group_id=current_group_id)
    
    entries = query.all()
    
    # Структуры для хранения занятости
    occupied_rooms = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    occupied_teachers = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    
    for entry in entries:
        week_key = entry.week_type
        day_key = str(entry.weekday)
        time_key = str(entry.time_slot)
        
        # Добавляем информацию о занятом кабинете
        if entry.room:
            occupied_rooms[week_key][day_key][time_key].append({
                'id': str(entry.room.id),
                'name': str(entry.room.number),
                'group': entry.group.name
            })
        
        # Добавляем информацию о занятом преподавателе
        if entry.teacher:
            occupied_teachers[week_key][day_key][time_key].append({
                'id': str(entry.teacher.id),
                'name': entry.teacher.get_full_name(),
                'group': entry.group.name
            })
    
    return dict(occupied_rooms), dict(occupied_teachers)


def check_room_conflict(week_type, weekday, time_slot, room_id, current_group_id=None):
    """
    Проверяет конфликт кабинета в указанное время.
    
    Returns:
        str or None: Сообщение о конфликте или None, если конфликта нет
    """
    if not room_id:
        return None
    
    conflict_query = ScheduleEntry.objects.filter(
        week_type=week_type,
        weekday=weekday,
        time_slot=time_slot,
        room_id=room_id
    ).select_related('group', 'room')
    
    if current_group_id:
        conflict_query = conflict_query.exclude(group_id=current_group_id)
    
    conflict_entry = conflict_query.first()
    
    if conflict_entry:
        time_slot_name = get_time_slot_display(time_slot)
        weekday_name = get_weekday_display(weekday)
        week_name = 'Верхняя' if week_type == 'top' else 'Нижняя'
        
        return (f"{week_name} неделя, {weekday_name}, {time_slot_name}: "
                f"Кабинет {conflict_entry.room.number} уже занят группой {conflict_entry.group.name}")
    
    return None


def check_teacher_conflict(week_type, weekday, time_slot, teacher_id, current_group_id=None):
    """
    Проверяет конфликт преподавателя в указанное время.
    
    Returns:
        str or None: Сообщение о конфликте или None, если конфликта нет
    """
    if not teacher_id:
        return None
    
    conflict_query = ScheduleEntry.objects.filter(
        week_type=week_type,
        weekday=weekday,
        time_slot=time_slot,
        teacher_id=teacher_id
    ).select_related('group', 'teacher')
    
    if current_group_id:
        conflict_query = conflict_query.exclude(group_id=current_group_id)
    
    conflict_entry = conflict_query.first()
    
    if conflict_entry:
        time_slot_name = get_time_slot_display(time_slot)
        weekday_name = get_weekday_display(weekday)
        week_name = 'Верхняя' if week_type == 'top' else 'Нижняя'
        
        return (f"{week_name} неделя, {weekday_name}, {time_slot_name}: "
                f"Преподаватель {conflict_entry.teacher.get_full_name()} уже занят с группой {conflict_entry.group.name}")
    
    return None


def get_time_slot_display(time_slot_value):
    """
    Возвращает отображаемое название временного слота.
    """
    time_slots_dict = dict(ScheduleEntry.TIME_SLOT_CHOICES)
    return time_slots_dict.get(time_slot_value, f"Слот {time_slot_value}")


def get_weekday_display(weekday_value):
    """
    Возвращает отображаемое название дня недели.
    """
    weekdays_dict = dict(ScheduleEntry.WEEKDAY_CHOICES)
    return weekdays_dict.get(weekday_value, f"День {weekday_value}")


@login_required
@user_passes_test(is_scheduler)
def get_occupancy_api(request, group_id):
    """
    API endpoint для получения данных о занятости кабинетов и преподавателей.
    Используется для AJAX запросов.
    """
    try:
        occupied_rooms_data, occupied_teachers_data = get_occupancy_data(group_id)
        
        return JsonResponse({
            'success': True,
            'occupied_rooms': occupied_rooms_data,
            'occupied_teachers': occupied_teachers_data
        })
        
    except Exception as e:
        logger.error(f"Error getting occupancy data: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(is_scheduler)
def validate_schedule_slot(request):
    """
    API endpoint для валидации отдельного слота расписания.
    Используется для проверки конфликтов в реальном времени.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        week_type = data.get('week_type')
        weekday = int(data.get('weekday'))
        time_slot = int(data.get('time_slot'))
        room_id = data.get('room_id')
        teacher_id = data.get('teacher_id')
        group_id = data.get('group_id')
        
        conflicts = []
        
        # Проверяем конфликт кабинета
        if room_id:
            room_conflict = check_room_conflict(
                week_type, weekday, time_slot, room_id, group_id
            )
            if room_conflict:
                conflicts.append({
                    'type': 'room',
                    'message': room_conflict
                })
        
        # Проверяем конфликт преподавателя
        if teacher_id:
            teacher_conflict = check_teacher_conflict(
                week_type, weekday, time_slot, teacher_id, group_id
            )
            if teacher_conflict:
                conflicts.append({
                    'type': 'teacher',
                    'message': teacher_conflict
                })
        
        return JsonResponse({
            'success': True,
            'conflicts': conflicts,
            'has_conflicts': len(conflicts) > 0
        })
        
    except Exception as e:
        logger.error(f"Error validating schedule slot: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required  
@user_passes_test(is_scheduler)
def schedule_statistics(request, group_id):
    """
    Возвращает статистику по расписанию группы.
    """
    try:
        group = get_object_or_404(Group, id=group_id)
        
        # Подсчитываем статистику
        total_slots = len(ScheduleEntry.WEEKDAY_CHOICES) * len(ScheduleEntry.TIME_SLOT_CHOICES) * 2  # 2 недели
        filled_slots = ScheduleEntry.objects.filter(
            group=group,
            scheduler=request.user
        ).count()
        
        # Получаем данные о конфликтах
        occupied_rooms_data, occupied_teachers_data = get_occupancy_data(group_id)
        
        # Подсчитываем конфликты
        conflicts_count = 0
        entries = ScheduleEntry.objects.filter(group=group, scheduler=request.user)
        
        for entry in entries:
            if check_room_conflict(entry.week_type, entry.weekday, entry.time_slot, 
                                 entry.room.id, group_id):
                conflicts_count += 1
            if check_teacher_conflict(entry.week_type, entry.weekday, entry.time_slot, 
                                    entry.teacher.id, group_id):
                conflicts_count += 1
        
        return JsonResponse({
            'success': True,
            'statistics': {
                'total_slots': total_slots,
                'filled_slots': filled_slots,
                'free_slots': total_slots - filled_slots,
                'conflicts_count': conflicts_count,
                'completion_percentage': round((filled_slots / total_slots) * 100, 1) if total_slots > 0 else 0
            }
        })
        
    except Exception as e:
        logger.error(f"Error getting schedule statistics: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    


# views.py (fragment with password reset functions)
# views.py (fragment with password reset functions)
from django.contrib.sites.shortcuts import get_current_site
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import send_mail
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.conf import settings
import logging

logger = logging.getLogger(__name__)
CustomUser = get_user_model()











def password_reset_request(request):
    if request.method == "POST":
        email = request.POST.get('email')
        logger.debug(f"password_reset_request: email={email}")
        
        if not email:
            messages.error(request, 'Please enter your email address.')
            return render(request, 'registration/password_reset_form.html')

        try:
            user = CustomUser.objects.get(email__iexact=email, is_active=True)
            logger.debug(f"Найден пользователь: {user.email}")
        except CustomUser.DoesNotExist:
            logger.debug("Пользователь не найден или неактивен")
            messages.success(request, 'If an active account with that email exists, you will receive instructions to reset your password.')
            return redirect('password_reset_done')
        except CustomUser.MultipleObjectsReturned:
            logger.error(f"Найдено несколько активных пользователей с email {email}")
            messages.error(request, 'System error. Please contact the administrator.')
            return render(request, 'registration/password_reset_form.html')

        if not user.is_active:
            messages.error(request, 'Account is inactive. Please contact the administrator.')
            return render(request, 'registration/password_reset_form.html')
        
        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        logger.debug(f"Сгенерирован uid для {user.email}")
        
        try:
            current_site = get_current_site(request)
            site_name = current_site.name
            domain = current_site.domain
        except Exception as e:
            logger.error(f"Ошибка получения информации о сайте: {e}")
            site_name = getattr(settings, 'SITE_NAME', 'Our Site')
            domain = getattr(settings, 'DOMAIN_NAME', 'localhost:8000')

        logger.debug(f"Сайт: {site_name}, {domain}")

        from django.urls import reverse
        reset_path = reverse('password_reset_confirm', kwargs={'uidb64': uid, 'token': token})
        scheme = 'https' if request.is_secure() else 'http'
        reset_url = f"{scheme}://{domain}{reset_path}"
        logger.debug(f"Ссылка сброса сформирована для {user.email}")
        
        context = {
            'user': user,
            'site_name': site_name,
            'domain': domain,
            'scheme': scheme,
            'uid': uid,
            'token': token,
            'reset_url': reset_url,
        }
        
        try:
            from django.template.loader import render_to_string
            subject = f'Password Reset for {site_name}'
            text_message = render_to_string('registration/password_reset_email.txt', context)
            send_mail(
                subject=subject,
                message=text_message,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@example.com'),
                recipient_list=[user.email],
                fail_silently=False,
            )
            logger.info(f"Email сброса пароля отправлен на {user.email}")
        except Exception as e:
            logger.error(f"Ошибка отправки email сброса пароля на {user.email}: {e}", exc_info=True)
            messages.error(request, 'Error sending email. Please try again later or contact the administrator.')
            return render(request, 'registration/password_reset_form.html')
        
        messages.success(request, 'If an active account with that email exists, you will receive instructions to reset your password.')
        return redirect('password_reset_done')

    return render(request, 'registration/password_reset_form.html')











def password_reset_done(request):
    logger.debug("password_reset_done вызван")
    return render(request, 'registration/password_reset_done.html')


def password_reset_confirm(request, uidb64, token):
    logger.debug(f"password_reset_confirm вызван uid={uidb64}")
    
    user = None
    token_valid = False
    
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = CustomUser.objects.get(pk=uid)
        logger.debug(f"Найден пользователь по uid: {user.email}")
    except (TypeError, ValueError, OverflowError, CustomUser.DoesNotExist) as e:
        logger.warning(f"Ошибка декодирования uid или поиска пользователя: {e}")
        user = None
    except Exception as e:
        logger.error(f"Неожиданная ошибка в password_reset_confirm: {e}")
        user = None

    if user is not None:
        token_valid = default_token_generator.check_token(user, token)
        logger.debug(f"Проверка токена для {user.email}: {token_valid}")
        
        if token_valid:
            if request.method == 'POST':
                from django.contrib.auth.forms import SetPasswordForm
                form = SetPasswordForm(user, request.POST)
                if form.is_valid():
                    try:
                        form.save()
                        messages.success(request, 'Your password has been successfully changed.')
                        logger.info(f"Пароль успешно изменён для {user.email}")
                        return redirect('password_reset_complete')
                    except Exception as e:
                        logger.error(f"Ошибка сохранения нового пароля для {user.email}: {e}")
                        messages.error(request, 'Error saving password. Please try again.')
                        form = SetPasswordForm(user, request.POST)
                else:
                    logger.warning(f"Ошибки формы сброса пароля: {form.errors}")
                    messages.error(request, 'Please correct the errors below.')
            else:
                from django.contrib.auth.forms import SetPasswordForm
                form = SetPasswordForm(user)
            
            context = {
                'form': form,
                'validlink': True,
                'title': 'Enter new password'
            }
            return render(request, 'registration/password_reset_confirm.html', context)
    
    logger.warning("Токен недействителен или пользователь не найден")
    context = {
        'form': None,
        'validlink': False,
        'title': 'Invalid link'
    }
    messages.error(request, 'The password reset link is invalid. It may have already been used or has expired.')
    return render(request, 'registration/password_reset_confirm.html', context)


def password_reset_complete(request):
    logger.debug("password_reset_complete вызван")
    return render(request, 'registration/password_reset_complete.html')



# Вспомогательная функция для проверки роли
def is_tutor(user):
    return user.is_authenticated and hasattr(user, 'role') and user.role == CustomUser.TUTOR

# --- Views для тьютора ---

@login_required
@user_passes_test(is_tutor)
def tutor_create_student(request):
    """Представление для создания нового студента тьютором."""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        # Добавим проверку, что роль установлена в 'student'
        if form.is_valid():
            with transaction.atomic():
                # Создаем пользователя
                user = form.save(commit=False)
                user.role = CustomUser.STUDENT
                user.is_active = True
                user.save()
# StudentProfile создаётся автоматически через сигнал create_student_profile
# Достаточно убедиться что профиль существует:
                profile, _ = StudentProfile.objects.get_or_create(user=user)
                
                messages.success(request, f'Студент {user.get_full_name()} успешно создан.')
                # Перенаправляем на страницу группы или список студентов
                # (предполагаем, что тьютор работает с одной группой, или показываем список его групп)
                return redirect('tutor_dashboard') # Замените на ваш URL
                
    else:
        # Предзаполняем форму ролью студента
        form = CustomUserCreationForm(initial={'role': CustomUser.STUDENT})
        # Можно также скрыть поле роли в шаблоне или сделать его только для чтения
    
    return render(request, 'tutor/create_student.html', {'form': form})


@login_required
@user_passes_test(is_tutor)
def tutor_create_group(request):
    """Представление для создания новой группы тьютором."""
    if request.method == 'POST':
        group_name = request.POST.get('name', '').strip()
        if group_name:
            try:
                # Создаем группу
                group = Group.objects.create(name=group_name)
                
                # Привязываем группу к профилю тьютора
                tutor_profile, created = TutorProfile.objects.get_or_create(user=request.user)
                tutor_profile.groups.add(group)
                
                messages.success(request, f'Группа "{group.name}" успешно создана и назначена вам.')
                # Перенаправляем на страницу группы или список групп
                return redirect('tutor_group_detail', group_id=group.id) # Замените на ваш URL
                
            except Exception as e:
                messages.error(request, f'Ошибка при создании группы: {e}')
        else:
            messages.error(request, 'Имя группы не может быть пустым.')
    
    return render(request, 'tutor/create_group.html')




@login_required
@user_passes_test(is_tutor)
def tutor_manage_group_students(request, group_id):
    """Представление для управления студентами в группе (добавление/удаление) тьютором."""
    tutor_profile = get_object_or_404(TutorProfile, user=request.user)
    # Убедимся, что тьютор имеет доступ к этой группе
    group = get_object_or_404(tutor_profile.groups, id=group_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        student_id = request.POST.get('student_id')

        if action and student_id:
            try:
                student_profile = get_object_or_404(StudentProfile, id=student_id)
                
                # --- Логика для действия 'add' ---
                if action == 'add':
                    # Проверка 1: Студент уже в ЭТОЙ конкретной группе?
                    if student_profile in group.students.all():
                        # Сообщение об ошибке, если студент уже в группе
                        messages.error(
                            request, 
                            f"Ошибка: Студент {student_profile.user.get_full_name()} уже состоит в группе {group.name}. Добавление невозможно."
                        )
                    # Проверка 2: Студент уже состоит в ЛЮБОЙ из групп тьютора?
                    # Эта проверка решает проблему: студент не должен быть ни в одной группе тьютора
                    elif student_profile.groups.filter(id__in=tutor_profile.groups.all()).exists():
                         messages.error(
                            request, 
                            f"Ошибка: Студент {student_profile.user.get_full_name()} уже состоит в одной из ваших групп. Добавление в несколько групп невозможно."
                        )
                    else:
                        # Студент не состоит ни в одной из групп тьютора - можно добавлять
                        group.students.add(student_profile)
                        messages.success(
                            request, 
                            f"Студент {student_profile.user.get_full_name()} успешно добавлен в группу {group.name}."
                        )
                # --- Конец логики для 'add' ---
                
                # --- Логика для действия 'remove' ---
                elif action == 'remove':
                    # Проверим, состоит ли студент в группе
                    if student_profile not in group.students.all():
                         messages.warning(
                            request, 
                            f"Предупреждение: Студент {student_profile.user.get_full_name()} не состоит в группе {group.name}."
                        )
                    else:
                        group.students.remove(student_profile)
                        messages.success(
                            request, 
                            f"Студент {student_profile.user.get_full_name()} успешно удален из группы {group.name}."
                        )
                # --- Конец логики для 'remove' ---
                
                else:
                    messages.error(request, "Неверное действие.")
                    
            except StudentProfile.DoesNotExist:
                messages.error(request, "Ошибка: Студент не найден.")
            except Exception as e:
                 messages.error(request, f"Ошибка: {e}")
                 
        else:
            messages.error(request, "Ошибка: Не указаны действие или студент.")

        return redirect('tutor_manage_group_students', group_id=group.id)

    else: # GET request
        # Получаем данные для отображения
        try:
            tutor_groups = tutor_profile.groups.all()
            
            # --- ИСПРАВЛЕНА ЛОГИКА ---
            # Находим студентов, которые НЕ состоят НИ В ОДНОЙ из групп тьютора
            # Это студенты, которых тьютор может добавить в ЛЮБУЮ свою группу
            students_not_in_any_tutor_group = StudentProfile.objects.exclude(
                groups__in=tutor_groups
            ).select_related('user').order_by('user__last_name', 'user__first_name')
            
            # Студенты, уже состоящие в текущей редактируемой группе
            students_in_current_group = group.students.all().select_related('user')
            
            # Доступные студенты - это те, кто не в любой группе тьютора
            # (и, следовательно, не в текущей группе)
            available_students = students_not_in_any_tutor_group
            # --- КОНЕЦ ИСПРАВЛЕНИЯ ---

            context = {
                'group': group,
                'students_in_group': students_in_current_group,
                'available_students': available_students,
            }
            return render(request, 'tutor/manage_group_students.html', context)
        except TutorProfile.DoesNotExist:
            messages.error(request, 'Профиль тьютора не настроен.')
            raise PermissionDenied("Профиль тьютора не настроен.")

# ... (остальные view-функции) ...


from django.shortcuts import get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.contrib.auth import get_user_model
import secrets
import string

CustomUser = get_user_model()

@staff_member_required
def send_credentials_to_user(request, user_id):
    """
    Ручная отправка учетных данных пользователю (для админов)
    """
    user = get_object_or_404(CustomUser, id=user_id)
    
    if not user.email:
        messages.error(request, f'У пользователя {user.username} не указан email адрес.')
        return HttpResponseRedirect(reverse('admin:school_customuser_changelist'))
    
    try:
        from .utils.email_utils import send_user_credentials_email
        from .signals import _newly_created_users
    except ImportError:
        messages.error(request, 'Модуль отправки email не найден. Проверьте настройки.')
        return HttpResponseRedirect(reverse('admin:school_customuser_changelist'))
    
    # Определяем нужно ли генерировать новый пароль
    generate_new_password = not user.has_usable_password()
    password = None
    
    if generate_new_password:
        alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
        password = ''.join(secrets.choice(alphabet) for i in range(12))
        user.set_password(password)
        user.save(update_fields=['password'])
    
    # Отправляем email с паролем если он был сгенерирован или пользователь недавно создан
    force_send = generate_new_password or user.id in _newly_created_users
    success = send_user_credentials_email(user, password, force_send_password=force_send)
    
    # Убираем пользователя из списка новых
    _newly_created_users.discard(user.id)
    
    if success:
        if password:
            messages.success(request, f'Новый пароль сгенерирован и отправлен пользователю {user.username} на {user.email}')
        else:
            messages.success(request, f'Данные для входа отправлены пользователю {user.username} на {user.email}')
    else:
        messages.error(request, f'Ошибка при отправке email пользователю {user.username}')
    
    return HttpResponseRedirect(reverse('admin:school_customuser_changelist'))


@staff_member_required  
def reset_user_password(request, user_id):
    """
    Сброс пароля пользователя и отправка нового по email
    """
    user = get_object_or_404(CustomUser, id=user_id)
    
    if not user.email:
        messages.error(request, f'У пользователя {user.username} не указан email адрес.')
        return HttpResponseRedirect(reverse('admin:school_customuser_changelist'))
    
    try:
        from .utils.email_utils import send_password_reset_email
    except ImportError:
        messages.error(request, 'Модуль отправки email не найден. Проверьте настройки.')
        return HttpResponseRedirect(reverse('admin:school_customuser_changelist'))
    
    # Генерируем новый пароль
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
    new_password = ''.join(secrets.choice(alphabet) for i in range(12))
    user.set_password(new_password)
    user.save(update_fields=['password'])
    
    success = send_password_reset_email(user, new_password)
    
    if success:
        messages.success(request, f'Новый пароль отправлен пользователю {user.username} на {user.email}')
    else:
        messages.error(request, f'Ошибка при отправке нового пароля пользователю {user.username}')
    
    return HttpResponseRedirect(reverse('admin:school_customuser_changelist'))











@login_required
def faq_view(request):
    # Эта страница может быть доступна всем, или вы можете добавить проверку
    # if not request.user.is_authenticated:
    #     return redirect('login') # Раскомментируйте, если хотите, чтобы только авторизованные пользователи видели FAQ
    return render(request, 'school/faq.html') # Убедитесь, что путь к шаблону правильный
